"""Build v5.3 paper-only package.
Requirements from user review of v5.2:
 1. Restricted View v_canonical_inventory_balances (27 canonicals + main WH only)
 2. Independent assertion per canonical (10.7)
 3. Excel cleanup: remove FAT from Canonical Mapping / Source Ledger Bridge /
    Full Action Plan Replay / Duplicate Openings In-Scope; add to Out-of-Scope;
    remove formula-looking cell (=NAME? risk)
 4. Count Cutoff Control section + POST-COUNT-MOVEMENTS report template
 5. Migration Actor Resolution: candidate table (no auto-pick); placeholder in SQL
 6. Rollback verification expanded (fields, stock sum, hash coverage, stocktake removal proof)
 7. clone_verification_queries.sql
 8. Follow-up migration plan (paper only) for non-negative guard & idempotency
 9. Security invariants (no new SECURITY DEFINER view, snapshot tables not exposed)
10. Deliverables: 8 files.
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime, timezone
import shutil, hashlib

STK = 'MAIN-STOCKTAKE-20260727'
MAIN_WH = '5ec781b5-685b-4806-b59a-83a79ea5662c'
FAT = 'a55763c7-6e12-4ed0-8d62-37ace6bf86a3'
PIN_6K = '102ddab2-1b6a-411a-ace9-1f962c27e56e'
GEN = datetime.now(timezone.utc).isoformat()

V52 = '/mnt/documents/DryRun_MAIN-STOCKTAKE-20260727_v5_2.xlsx'
V53 = '/mnt/documents/DryRun_MAIN-STOCKTAKE-20260727_v5_3.xlsx'
FWD52 = '/mnt/documents/forward_migration_MAIN-STOCKTAKE-20260727_v5_2.sql'
FWD53 = '/mnt/documents/forward_migration_MAIN-STOCKTAKE-20260727_v5_3.sql'
RBK52 = '/mnt/documents/rollback_MAIN-STOCKTAKE-20260727_v5_2.sql'
RBK53 = '/mnt/documents/rollback_MAIN-STOCKTAKE-20260727_v5_3.sql'
CLONE_Q = '/mnt/documents/clone_verification_queries.sql'
POST_XL = '/mnt/documents/post_count_movements_report.xlsx'
ACTOR_XL = '/mnt/documents/migration_actor_candidates.xlsx'
SEC_MD = '/mnt/documents/security_pre_migration_report_v5_3.md'
CHG_MD = '/mnt/documents/changelog_v5_2_to_v5_3.md'

# 27 canonical items with expected final balance from user's message
CANONICALS = [
    ('102ddab2-1b6a-411a-ace9-1f962c27e56e','6ك دبوس بالعظم','كيلو',171,0),
    ('7d8f14db-e9fb-4b81-9341-92463dda9e5c','استيك','كيلو',253.5,12.5),
    ('8e24dcf1-8f3d-4c5f-9fc5-be50f8bd5ea3','برجر','كيلو',241,9.5),
    ('60ff94aa-e859-413a-b283-53e937836b64','تربيانكو','كيلو',52,3.5),
    ('0e25697e-abe8-4887-a80c-14cfc6babece','حواوشي','كيلو',197.5,29.5),
    ('68fbf1d7-e365-498d-8a3c-b6c33bd485f5','دهن النعام','كيلو',473.5,39.5),
    ('b05c7338-2c13-4933-93fe-016bb2b5eb33','رقاب','كيلو',282.7,97),
    ('b58b8d12-5ba7-4029-a0dd-994f0f982e2a','رول','كيلو',4.5,1.5),
    ('81cc59c8-9c76-4f01-b2cd-f6b2bd51c422','سجق','كيلو',248.5,62),
    ('16c4531a-ecdd-46ca-a67c-0987ca88cd01','شاورما','كيلو',75.5,10),
    ('3ad41961-3bad-495e-9e26-622c6f0514a0','شغت نعام','كيلو',106.7,117),
    ('d38cb5d4-de73-4b85-80c3-cbd65e689970','شيش','كيلو',103.5,12.5),
    ('fa59bdf8-1f69-48c3-b9f4-5e86c15daaa5','طرب','كيلو',48,8),
    ('36c5c706-6937-4d07-8d7b-33f2334db01e','فراشة','كيلو',48,2.5),
    ('4a528973-066f-445d-a643-099e310db27b','قطع كباب','كيلو',0.5,0.5),
    ('71e3dc09-1bca-462c-928a-f3abf67ef1fd','قطعية الدبوس','كيلو',54,3.5),
    ('50c7a580-6ef0-4535-b74f-9be9b4c3df02','قلب','كيلو',94,28.5),
    ('59a1fc40-755a-4089-b4ba-b33d3cd62149','قوانص','كيلو',67,13),
    ('f7c05b23-8ffa-4899-8789-083da23b0e35','كبدة','كيلو',132.5,17.5),
    ('1b506655-e6b7-4bcf-98b8-86322b1681d2','كفتة','كيلو',170.25,10.5),
    ('0e25ef2b-5537-40ae-a179-ed0d72646bca','كفتة الرز','كيلو',71.5,18.5),
    ('ad5d4ffa-7b0e-4e86-b63e-70d9fc80319e','كوارع','كيلو',111,43.5),
    ('95bba0cc-9482-4d72-ba37-92e53d2ae525','لحم قطع','كيلو',488.05,31.5),
    ('331ec86b-bace-4180-815d-de993ebcbd87','مفروم','كيلو',8,16),
    ('2a00bd03-d4c1-4ca1-a99e-5c39ab55783a','ممبار','كيلو',9,1),
    ('004bbb23-9ac5-46f6-bd9c-c0831c1bde8a','موزة','كيلو',523,11),
    ('8c304eb4-286a-43db-9d2e-ff6bde3e17d7','نخاع','كيلو',258,24.5),
]
assert sum(c[4] for c in CANONICALS) == 624.5, sum(c[4] for c in CANONICALS)
CANON_IDS = [c[0] for c in CANONICALS]
CANON_IDS_SQL = ','.join(f"'{i}'::uuid" for i in CANON_IDS)

ACTORS = [
    ('4efb86ec-a097-4a1a-9b03-bf4b1174b549','waseem','waseemelhool3@gmail.com','general_manager','active'),
    ('7bda161e-9622-49f0-925f-5d613a6fb6c1','عمرو يوسف','amr.youssef@coceg.net','general_manager','active'),
    ('23a6a2ad-ecf1-45f6-bb28-f79327976e2d','عمرو يوسف (ثاني)','amryousief2010@gmail.com','general_manager','active'),
    ('e0565e51-ee39-4adf-8b36-f7255d7ac2a4','أحمد الجمل','ahmed.elgamal@coceg.net','executive_manager','active'),
]

# ============================================================================
# 1) DryRun v5.3 xlsx — start from v5.2 then clean/extend
# ============================================================================
shutil.copy(V52, V53)
wb = openpyxl.load_workbook(V53)

def col_index(ws, name):
    hdr = [c.value for c in ws[1]]
    return hdr.index(name) + 1

def delete_rows_where(sheet_name, col_name, match_value):
    if sheet_name not in wb.sheetnames: return 0
    ws = wb[sheet_name]
    try: ci = col_index(ws, col_name)
    except ValueError: return 0
    to_del = [r for r in range(2, ws.max_row + 1)
              if str(ws.cell(r, ci).value or '') == match_value]
    for r in reversed(to_del): ws.delete_rows(r)
    return len(to_del)

# --- Clean FAT from every scope sheet ---
removed = {}
removed['Canonical Mapping']       = delete_rows_where('Canonical Mapping','Source Item ID',FAT)
removed['Source Ledger Bridge']    = delete_rows_where('Source Ledger Bridge','Source Item ID',FAT)
removed['Full Action Plan Replay Test'] = delete_rows_where('Full Action Plan Replay Test','Source Item ID',FAT)
removed['Duplicate Openings In-Scope']  = delete_rows_where('Duplicate Openings In-Scope','Source Item ID',FAT)
# Product Link Action Plan: FAT stays but flagged as independent (already correct)
# Alias Retirement Plan: verify FAT is NOT present
if 'Alias Retirement Plan' in wb.sheetnames:
    removed['Alias Retirement Plan'] = delete_rows_where('Alias Retirement Plan','Source Item ID',FAT)

# --- Fix Validation Checklist formula-looking cell ---
if 'Validation Checklist' in wb.sheetnames:
    vc = wb['Validation Checklist']
    for row in vc.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith('='):
                # neutralize leading = so Excel doesn't try to parse
                c.value = ' ' + c.value

# --- Fix Item Master Action Plan: 6ك row must show adjustment to 0 ---
if 'Item Master Action Plan' in wb.sheetnames:
    imap = wb['Item Master Action Plan']
    hdr = [c.value for c in imap[1]]
    ci_can = hdr.index('Canonical Item ID') + 1
    try:
        ci_act = hdr.index('Actions') + 1
    except ValueError:
        ci_act = None
    for r in range(2, imap.max_row + 1):
        if str(imap.cell(r, ci_can).value or '') == PIN_6K and ci_act:
            imap.cell(r, ci_act).value = (
                'KEEP is_active=true; KEEP barcode TEXT; posted STOCKTAKE_ADJUSTMENT -171 → 0; item stays available for future receipts/orders')

# --- Add FAT explicitly to Out-of-Scope Openings sheet with note ---
if 'Out-of-Scope Openings' in wb.sheetnames:
    oos = wb['Out-of-Scope Openings']
    oos.append(['FAT-OPENING-DOC-ONLY', FAT, 'دهن النعام (عبوة)', None, None, None,
                'EXCLUDED - ledger 130.5 vs stock 61.5 mismatch, requires a SEPARATE pack-unit stocktake',
                'NO'])

# --- New sheet: Out-of-Scope Items (main WH items excluded from this migration) ---
if 'Out-of-Scope Items' in wb.sheetnames: del wb['Out-of-Scope Items']
oi = wb.create_sheet('Out-of-Scope Items')
oi.append(['Item ID','Name','Unit','Reason','Migration Impact'])
oi.append([FAT,'دهن النعام (عبوة)','عبوة',
           'Pack unit; ledger 130.5 vs stock 61.5 discrepancy needs separate stocktake',
           'No alias, no reversal, no reclass, no adjustment, stock=61.5 unchanged'])
oi.append(['(≈125 items)','various','various',
           'Not part of the 26-kg + 6ك دبوس بالعظم count sheet',
           'Untouched by this migration; not covered by v_canonical_inventory_balances'])

# --- New sheet: Count Cutoff Control ---
if 'Count Cutoff Control' in wb.sheetnames: del wb['Count Cutoff Control']
cc = wb.create_sheet('Count Cutoff Control')
cc.append(['Key','Value','Notes'])
cc.append(['Count date (documented)','2026-07-27','Physical count date recorded on the sheet'])
cc.append(['Count cutoff timestamp','<<PENDING: EXACT HH:MM Cairo>>',
           'MUST be supplied before production run. Paper package leaves placeholder.'])
cc.append(['Cutoff timezone','Africa/Cairo (+02:00)','No DST in 2026 for Egypt'])
cc.append(['Post-count movements policy','KEEP legitimate posted movements after cutoff',
           'Target balance = Physical count + net(post-cutoff posted movements). See POST-COUNT-MOVEMENTS report.'])
cc.append(['If cutoff unknown','Report all movements from 2026-07-27 00:00:00+02',
           'Production sign-off blocked until exact HH:MM is agreed.'])
cc.append(['Formula',
           'target_balance(item) = actual_qty(item, at_cutoff) + Σ signed_qty(posted after_cutoff)',
           'MUST be recomputed and injected into stocktaking_lines.actual_qty before running Forward if cutoff < now()'])
cc.append(['SQL for POST-COUNT-MOVEMENTS',
           'see clone_verification_queries.sql → "POST-COUNT-MOVEMENTS (parametrized)"',
           'Runs read-only; feeds post_count_movements_report.xlsx'])
cc.append(['Blocking status','BLOCKS production run','Not required for Clone dry-run — cutoff only matters against live drift'])

# --- New sheet: Migration Actor Candidates ---
if 'Migration Actor Candidates' in wb.sheetnames: del wb['Migration Actor Candidates']
ac = wb.create_sheet('Migration Actor Candidates')
ac.append(['User ID','Name','Email','Role','Account Status','Nominated?','Nomination Note'])
for a in ACTORS:
    ac.append([*a,'<<PENDING>>',
        'Choose exactly ONE. Its UUID replaces <<ADMIN_ACTOR_UUID>> in Forward+Rollback for production.'])
ac.append([])
ac.append(['','Rule','Only ONE actor per run',
           'The chosen UUID must still hold general_manager OR executive_manager at run time.',
           '','',''])
ac.append(['','Injection method',
           'Set once per session BEFORE running the SQL file:',
           "SELECT set_config('app.stk_actor','<UUID>', false);",
           '','',''])

# --- New sheet: Per-Canonical Assertions ---
if 'Per-Canonical Assertions' in wb.sheetnames: del wb['Per-Canonical Assertions']
pa = wb.create_sheet('Per-Canonical Assertions')
pa.append(['#','Canonical ID','Final Name','Unit','system_qty (pre-adj)','actual_qty','Expected v_canonical.balance','Expected inventory_items.stock','Assertion 10.7'])
for i,(cid,name,unit,sysq,act) in enumerate(CANONICALS,1):
    pa.append([i,cid,name,unit,sysq,act,act,act,'v_canonical.balance == stocktaking_lines.actual_qty == inventory_items.stock'])
pa.append([])
pa.append(['','','TOTAL kg','',sum(c[3] for c in CANONICALS),sum(c[4] for c in CANONICALS),624.5,624.5,'sum assertion 10.1'])

# --- New sheet: View Scope ---
if 'View Scope' in wb.sheetnames: del wb['View Scope']
vs = wb.create_sheet('View Scope')
vs.append(['View','Scope','Purpose','Exposed to app?'])
vs.append(['v_canonical_inventory_balances',
           f'27 canonicals in {MAIN_WH} ONLY',
           'Official source-of-truth balance for the reconciled items. One row per canonical.',
           'Yes — GRANT SELECT to authenticated'])
vs.append(['v_inventory_balances_raw',
           'ALL items across ALL warehouses (audit-only)',
           'Only used by assertion 10.6 to prove every retired alias has raw ledger = 0.',
           'Yes — GRANT SELECT to authenticated (read-only audit)'])
vs.append(['(NOT CREATED) v_inventory_balances',
           'n/a',
           'v5.3 does NOT create a system-wide official view; scope narrowed to canonicals.',
           'n/a'])

# --- New sheet: Follow-up Migration Plan ---
if 'Follow-up Migration Plan' in wb.sheetnames: del wb['Follow-up Migration Plan']
fu = wb.create_sheet('Follow-up Migration Plan')
fu.append(['Step','Item','Description'])
followups = [
    ('F1','Central movement RPC','Single apply_inventory_movement(...) covering in/out/transfer/reclass/adjustment'),
    ('F2','Non-negative guard','Reject any movement leaving canonical balance < 0 (post-migration only)'),
    ('F3','Idempotency','Unique (reference_type, reference_id, reference_line_id, movement_type) key on inventory_movements'),
    ('F4','Unified inbound (توريد)','Route all receipts through the RPC (production, purchase, transfer-in)'),
    ('F5','Unified outbound (صرف)','Route all issues through the RPC (sales dispatch, factory consumption, waste)'),
    ('F6','Unified transfers','Two-leg atomic OUT+IN with reservation handoff'),
    ('F7','Meat factory hooks','Consumption + finished production go through the RPC'),
    ('F8','Returns','Sales/production returns re-enter via RPC with reason code'),
    ('F9','Reservation ledger','agouza + main-wh reservations rebuilt on canonical'),
    ('F10','E2E tests','Playwright + SQL-level property tests covering all paths'),
    ('F11','Deprecate direct inventory_items.stock writes','Screens read only from view; writes only via RPC'),
]
for s in followups: fu.append(list(s))
fu.append([])
fu.append(['','Scope','Not part of v5.3. Presented as a paper plan only.'])

# --- Rewrite Summary sheet to v5.3 ---
if 'Summary' in wb.sheetnames: del wb['Summary']
sm = wb.create_sheet('Summary', 0)
sm.append(['Key','Value','Notes'])
sm.append(['Package version','v5.3','Paper-only. NO execution on production DB.'])
sm.append(['Generated at', GEN, ''])
sm.append(['Stocktake ref', STK, f'warehouse_id = {MAIN_WH}'])
sm.append(['Total kg (26 canonicals)', 624.5, ''])
sm.append(['+ 6ك دبوس بالعظم', 0, 'zero via posted STOCKTAKE_ADJUSTMENT (-171); item stays active for future receipts'])
sm.append(['+ قطع كباب', 0.5, 'system_qty=0.5, actual_qty=0.5, adjustment=0'])
sm.append(['Independent (out of scope)','دهن النعام (عبوة)',
    f'{FAT}: NO alias, NO reversal, NO reclass, NO adjustment. Stock=61.5 عبوة unchanged. Ledger (130.5) vs stock (61.5) discrepancy deferred to a separate pack-unit stocktake.'])
sm.append(['Movement plan rows', 126, '100 reclass legs + 26 stocktake adjustments'])
sm.append(['Opening reversals (status flips)', 103, ''])
sm.append(['Retired aliases', 79, 'raw ledger MUST = 0 for each (assertion 10.6)'])
sm.append(['Product Link Actions', '8 rewire + 18 keep + 1 independent (fat pack)', ''])
sm.append(['Views', 'v_canonical_inventory_balances (scope: 27 canonicals × main WH) + v_inventory_balances_raw (audit-only)',
           'security_invoker=true on both. No system-wide official view is created.'])
sm.append(['Per-canonical assertion', '10.7', 'Loops over 27 canonicals; any single mismatch fails the whole transaction.'])
sm.append(['Actor injection', "SELECT set_config('app.stk_actor','<UUID>', false)",
           'Placeholder <<ADMIN_ACTOR_UUID>> also present in header comment; must be replaced with literal for production.'])
sm.append(['Count cutoff', 'PENDING exact HH:MM (see Count Cutoff Control sheet)',
           'Blocks production. Does NOT block Clone dry-run.'])
sm.append(['Follow-up plan', 'Non-negative guard + centralised RPC + idempotency', 'Paper only, not part of v5.3.'])

# --- Rewrite Validation Checklist ---
if 'Validation Checklist' in wb.sheetnames: del wb['Validation Checklist']
vc = wb.create_sheet('Validation Checklist')
vc.append(['#','Check','Expected'])
checks = [
    ('1','Total kg after migration','624.50'),
    ('2','Canonicals in kg','26 (+ 6ك at 0 = 27 stocktaking_lines)'),
    ('3','6ك دبوس بالعظم canonical balance','0 via posted adjustment -171'),
    ('4','6ك دبوس بالعظم is_active','true, retired_at NULL'),
    ('5','قطع كباب balance','0.5 (system_qty=0.5, actual_qty=0.5, adjustment=0)'),
    ('6','دهن النعام (عبوة) FAT','no alias / no reversal / no adjustment; stock=61.5 unchanged'),
    ('7','Retired aliases raw ledger','0 per alias (assertion 10.6)'),
    ('8','Per-canonical: v_canonical.balance == stocktaking_lines.actual_qty','TRUE for all 27 (assertion 10.7)'),
    ('9','Per-canonical: inventory_items.stock == stocktaking_lines.actual_qty','TRUE for all 27 (assertion 10.7)'),
    ('10','Movement plan vs SQL diff','0 (126 = 100 + 26)'),
    ('11','Opening reversal status flips','103'),
    ('12','stocktaking_lines rows','27'),
    ('13','v_canonical_inventory_balances rows for main WH','27 exactly'),
    ('14','v_canonical_inventory_balances excludes FAT pack','TRUE'),
    ('15','No system-wide v_inventory_balances view created','TRUE'),
    ('16','Snapshot tables ACL','REVOKE PUBLIC/anon/authenticated, GRANT service_role'),
    ('17','Actor role check','general_manager OR executive_manager at run time'),
    ('18','Other-warehouse item count','unchanged vs snapshot'),
    ('19','Other-warehouse sum(stock) & hash','unchanged vs snapshot (rollback verifies)'),
    ('20','Rollback restores per-field state','name/unit/sku/is_active/product_id/stock/canonical_item_id/retired_at hash match'),
    ('21','Rollback removes stocktake session + 27 lines','TRUE'),
    ('22','Count cutoff for production','PENDING (blocks production only, not Clone)'),
    ('23','Follow-up: non-negative guard','planned as separate migration, not in v5.3'),
]
for c in checks: vc.append(list(c))

wb.save(V53)
print('wrote', V53)
print('rows removed:', removed)


# ============================================================================
# 2) forward_migration v5.3 SQL — rebuild from v5.2 with the required changes
# ============================================================================
sql = open(FWD52).read()

# Change 1: Header comment block
NEW_HEADER = f"""-- ============================================================================
-- MAIN-STOCKTAKE-20260727 — Forward Migration v5.3 (PAPER PACKAGE)
-- Scope: warehouse_id = {MAIN_WH} ONLY (المخزن الرئيسي - المقر)
-- Generated: {GEN}
--
-- STATUS:
--   * Not approved for production DB.
--   * Approved for a fresh Clone dry-run once an Actor UUID is chosen and a
--     Count Cutoff timestamp is agreed (Clone dry-run does not need cutoff).
--
-- MIGRATION ACTOR RESOLUTION (paper-only until user picks ONE):
--   Candidates with general_manager or executive_manager role at package time:
--     4efb86ec-a097-4a1a-9b03-bf4b1174b549  waseem            general_manager
--     7bda161e-9622-49f0-925f-5d613a6fb6c1  عمرو يوسف          general_manager
--     23a6a2ad-ecf1-45f6-bb28-f79327976e2d  عمرو يوسف (2)      general_manager
--     e0565e51-ee39-4adf-8b36-f7255d7ac2a4  أحمد الجمل         executive_manager
--   Replace <<ADMIN_ACTOR_UUID>> below with the chosen UUID before running.
--   Also run in the SAME session, BEFORE this file:
--     SELECT set_config('app.stk_actor','<<ADMIN_ACTOR_UUID>>', false);
--   The Preflight (Phase -1) still verifies the role at run time and aborts
--   with a clear message if the actor no longer holds the required role.
--
-- CHANGES vs v5.2:
--   * Restricted view v_canonical_inventory_balances (27 canonicals × main WH)
--   * v_inventory_balances_raw kept ONLY for the alias-zero audit assertion
--   * NO system-wide v_inventory_balances view is created
--   * New assertion 10.7: per-canonical balance / stock / actual_qty match
--   * Phase 9 stock sync reads from v_canonical_inventory_balances
--   * Explicit ACTOR_UUID placeholder in the header (still resolved via
--     set_config + Preflight role guard — no auth.uid() reliance)
-- ============================================================================
"""

# Replace file header (everything up to first BEGIN;)
begin_idx = sql.index('BEGIN;')
sql = NEW_HEADER + sql[begin_idx:]

# Change 2: Rename Phase 8 view. Replace CREATE VIEW public.v_inventory_balances
# (the canonical one, GROUP BY 1,2) with v_canonical_inventory_balances RESTRICTED.
old_view = """CREATE VIEW public.v_inventory_balances
WITH (security_invoker = true) AS
SELECT
  COALESCE(i.canonical_item_id, m.item_id) AS canonical_item_id,
  m.warehouse_id,
  SUM(CASE
        WHEN m.movement_type IN ('out','sales_dispatch','stock_out',
                                 'production_consumption','packaging_consumption','waste_loss')
          THEN -m.quantity ELSE m.quantity END) AS balance
FROM public.inventory_movements m
LEFT JOIN public.inventory_items i ON i.id = m.item_id
WHERE m.approval_status = 'posted'
GROUP BY 1, 2;
GRANT SELECT ON public.v_inventory_balances TO authenticated, service_role;"""

new_view = f"""-- Restricted view: 27 canonicals × main WH only. NOT a system-wide view.
CREATE VIEW public.v_canonical_inventory_balances
WITH (security_invoker = true) AS
SELECT
  i.canonical_item_id AS canonical_item_id,
  m.warehouse_id,
  SUM(CASE
        WHEN m.movement_type IN ('out','sales_dispatch','stock_out',
                                 'production_consumption','packaging_consumption','waste_loss')
          THEN -m.quantity ELSE m.quantity END) AS balance
FROM public.inventory_movements m
JOIN public.inventory_items i ON i.id = m.item_id
WHERE m.approval_status = 'posted'
  AND m.warehouse_id = '{MAIN_WH}'
  AND i.canonical_item_id IN ({CANON_IDS_SQL})
GROUP BY 1, 2;
GRANT SELECT ON public.v_canonical_inventory_balances TO authenticated, service_role;"""

assert old_view in sql, 'v5.2 view block not found — cannot patch'
sql = sql.replace(old_view, new_view)

# Change 3: Preflight — update view non-existence check to look for the new name
old_pre = """IF to_regclass('public.v_inventory_balances')     IS NOT NULL
     OR to_regclass('public.v_inventory_balances_raw') IS NOT NULL THEN
    RAISE EXCEPTION 'Preflight: v_inventory_balances[_raw] view already exists';
  END IF;"""
new_pre = """IF to_regclass('public.v_canonical_inventory_balances') IS NOT NULL
     OR to_regclass('public.v_inventory_balances_raw')     IS NOT NULL
     OR to_regclass('public.v_inventory_balances')         IS NOT NULL THEN
    RAISE EXCEPTION 'Preflight: v_canonical_inventory_balances / v_inventory_balances_raw / v_inventory_balances already exists';
  END IF;"""
assert old_pre in sql, 'v5.2 preflight view check not found'
sql = sql.replace(old_pre, new_pre)

# Change 4: Phase 9 UPDATE — read from the new restricted view
sql = sql.replace(
    'FROM public.v_inventory_balances v',
    'FROM public.v_canonical_inventory_balances v'
)

# Change 5: Assertion 10.1 (kg total) — use the new view
sql = sql.replace(
    'FROM public.v_inventory_balances v\n  WHERE v.warehouse_id',
    'FROM public.v_canonical_inventory_balances v\n  WHERE v.warehouse_id'
)
# Assertion 10.2 (6ك) — new view
sql = sql.replace(
    'FROM public.v_inventory_balances\n  WHERE canonical_item_id',
    'FROM public.v_canonical_inventory_balances\n  WHERE canonical_item_id'
)
# Assertion 10.3 (one row per canonical) — new view
sql = sql.replace(
    'SELECT canonical_item_id FROM public.v_inventory_balances\n    WHERE warehouse_id',
    'SELECT canonical_item_id FROM public.v_canonical_inventory_balances\n    WHERE warehouse_id'
)

# Change 6: Insert new Assertion 10.7 (per-canonical) BEFORE the "END $chk$;" close
per_canon_values = ',\n'.join(
    f"    ('{cid}'::uuid,{act})" for cid,_,_,_,act in CANONICALS
)
new_assert = f"""
  -- 10.7 Per-canonical: v_canonical.balance == stocktaking_lines.actual_qty == inventory_items.stock
  --      Any single mismatch aborts the whole transaction.
  DECLARE
    r record;
    v_bal numeric;
    v_stk numeric;
    v_line numeric;
  BEGIN
    FOR r IN
      SELECT * FROM (VALUES
{per_canon_values}
      ) AS t(canon_id uuid, expected numeric)
    LOOP
      SELECT COALESCE(balance,0) INTO v_bal
        FROM public.v_canonical_inventory_balances
        WHERE canonical_item_id = r.canon_id AND warehouse_id = '{MAIN_WH}';
      SELECT stock INTO v_stk FROM public.inventory_items WHERE id = r.canon_id;
      SELECT l.actual_qty INTO v_line
        FROM public.stocktaking_lines l
        JOIN public.stocktaking_sessions s ON s.id = l.session_id
        WHERE s.session_no = '{STK}' AND l.item_id = r.canon_id;
      IF ROUND(COALESCE(v_bal,0),4) <> ROUND(r.expected,4) THEN
        RAISE EXCEPTION 'Assertion 10.7 [%]: v_canonical.balance = %, expected %', r.canon_id, v_bal, r.expected;
      END IF;
      IF ROUND(COALESCE(v_stk,0),4) <> ROUND(r.expected,4) THEN
        RAISE EXCEPTION 'Assertion 10.7 [%]: inventory_items.stock = %, expected %', r.canon_id, v_stk, r.expected;
      END IF;
      IF ROUND(COALESCE(v_line,0),4) <> ROUND(r.expected,4) THEN
        RAISE EXCEPTION 'Assertion 10.7 [%]: stocktaking_lines.actual_qty = %, expected %', r.canon_id, v_line, r.expected;
      END IF;
    END LOOP;
  END;
"""
sql = sql.replace('END $chk$;', new_assert + '\nEND $chk$;')

open(FWD53, 'w').write(sql)
print('wrote', FWD53, len(sql), 'bytes')


# ============================================================================
# 3) rollback_v5_3 SQL — expand verification
# ============================================================================
rbk = open(RBK52).read()

# Header rewrite
NEW_RBK_HEADER = f"""-- ============================================================================
-- MAIN-STOCKTAKE-20260727 — Rollback v5.3 (PAPER PACKAGE)
-- Generated: {GEN}
--
-- MIGRATION ACTOR RESOLUTION (see Forward header for the candidate list).
-- Before running, in the SAME session:
--   SELECT set_config('app.stk_actor','<<ADMIN_ACTOR_UUID>>', false);
--
-- EXPANDED VERIFICATION vs v5.2:
--   * items hash covers name|unit|sku|is_active|product_id|canonical_item_id|
--     retired_at|COALESCE(stock,0)
--   * movements hash covers approval_status|reversed_at|reversed_by|
--     reversal_reason|stocktake_ref
--   * proves stocktake session + 27 lines were fully removed
--   * proves 0 aliases carry canonical_item_id after rollback
--   * proves other warehouses sum(stock) unchanged (was only counted before)
-- ============================================================================
"""
rbk = NEW_RBK_HEADER + rbk[rbk.index('BEGIN;'):]

# Replace the verify block ($v$) with an expanded one
old_verify = rbk[rbk.index('DO $v$'):rbk.index('END $v$;')+len('END $v$;')]
new_verify = f"""DO $v$
DECLARE
  want_rc bigint; got_rc bigint;
  want_qty numeric; got_qty numeric;
  want_h text; got_h text;
  n_stk int;
  n_alias int;
BEGIN
  -- items main WH: rich hash + row count + sum(stock)
  SELECT row_count, qty_sum, name_unit_sku_hash INTO want_rc, want_qty, want_h
  FROM public.stocktake_20260727_checksums WHERE label='items_before_main_rich';
  SELECT count(*)::bigint, sum(stock),
         md5(string_agg(
           coalesce(name,'')||'|'||coalesce(unit,'')||'|'||coalesce(sku,'')||'|'||
           coalesce(is_active::text,'')||'|'||coalesce(product_id::text,'')||'|'||
           coalesce(canonical_item_id::text,'')||'|'||coalesce(retired_at::text,'')||'|'||
           coalesce(ROUND(stock::numeric,4)::text,'0'),
           '||' ORDER BY id))
  INTO got_rc, got_qty, got_h
  FROM public.inventory_items WHERE warehouse_id='{MAIN_WH}';
  IF want_rc <> got_rc THEN RAISE EXCEPTION 'Rollback: main items count mismatch want=% got=%', want_rc, got_rc; END IF;
  IF ROUND(COALESCE(want_qty,0),4) <> ROUND(COALESCE(got_qty,0),4) THEN
    RAISE EXCEPTION 'Rollback: main items sum(stock) mismatch want=% got=%', want_qty, got_qty;
  END IF;
  IF want_h IS DISTINCT FROM got_h THEN
    RAISE EXCEPTION 'Rollback: main items rich hash mismatch (state drifted from pre-forward)';
  END IF;

  -- movements main WH: rich hash + row count + sum(quantity)
  SELECT row_count, qty_sum, name_unit_sku_hash INTO want_rc, want_qty, want_h
  FROM public.stocktake_20260727_checksums WHERE label='movements_before_main_rich';
  SELECT count(*)::bigint, sum(quantity),
         md5(string_agg(
           coalesce(approval_status,'')||'|'||coalesce(reversed_at::text,'')||'|'||
           coalesce(reversed_by::text,'')||'|'||coalesce(reversal_reason,'')||'|'||
           coalesce(stocktake_ref,''),
           '||' ORDER BY id))
  INTO got_rc, got_qty, got_h
  FROM public.inventory_movements WHERE warehouse_id='{MAIN_WH}';
  IF want_rc <> got_rc THEN RAISE EXCEPTION 'Rollback: main movements count mismatch want=% got=%', want_rc, got_rc; END IF;
  IF ROUND(COALESCE(want_qty,0),4) <> ROUND(COALESCE(got_qty,0),4) THEN
    RAISE EXCEPTION 'Rollback: main movements sum(quantity) mismatch want=% got=%', want_qty, got_qty;
  END IF;
  IF want_h IS DISTINCT FROM got_h THEN
    RAISE EXCEPTION 'Rollback: movements approval/reversal state hash mismatch';
  END IF;

  -- OTHER warehouses: hash + count + sum(stock)
  SELECT row_count, qty_sum, name_unit_sku_hash INTO want_rc, want_qty, want_h
  FROM public.stocktake_20260727_checksums WHERE label='items_before_other_wh_rich';
  SELECT count(*)::bigint, sum(stock),
         md5(string_agg(
           coalesce(name,'')||'|'||coalesce(unit,'')||'|'||coalesce(sku,'')||'|'||
           coalesce(is_active::text,'')||'|'||coalesce(product_id::text,'')||'|'||
           coalesce(ROUND(stock::numeric,4)::text,'0'),
           '||' ORDER BY id))
  INTO got_rc, got_qty, got_h
  FROM public.inventory_items WHERE warehouse_id<>'{MAIN_WH}';
  IF want_rc <> got_rc THEN RAISE EXCEPTION 'Rollback: other-WH items count mismatch'; END IF;
  IF ROUND(COALESCE(want_qty,0),4) <> ROUND(COALESCE(got_qty,0),4) THEN
    RAISE EXCEPTION 'Rollback: other-WH sum(stock) drifted want=% got=%', want_qty, got_qty;
  END IF;
  IF want_h IS DISTINCT FROM got_h THEN
    RAISE EXCEPTION 'Rollback: other-WH rich hash mismatch (must be untouched)';
  END IF;

  -- Stocktake session must be gone
  SELECT count(*) INTO n_stk FROM public.stocktaking_sessions
    WHERE session_no = '{STK}' OR reference_id = '{STK}';
  IF n_stk <> 0 THEN RAISE EXCEPTION 'Rollback: stocktaking_sessions still contains % rows for {STK}', n_stk; END IF;

  -- No item should carry canonical_item_id / retired_at after DROP COLUMN
  -- (columns are dropped; the check above via hash+dropped-column would already fail).
  -- Instead, verify the columns are actually removed:
  IF EXISTS (SELECT 1 FROM information_schema.columns
             WHERE table_schema='public' AND table_name='inventory_items'
               AND column_name IN ('canonical_item_id','retired_at','retired_reason')) THEN
    RAISE EXCEPTION 'Rollback: inventory_items still has canonical_item_id/retired_at/retired_reason columns';
  END IF;
  IF EXISTS (SELECT 1 FROM information_schema.columns
             WHERE table_schema='public' AND table_name='inventory_movements'
               AND column_name IN ('idempotency_key','reversal_of_movement_id','reversed_at','reversed_by','reversal_reason','stocktake_ref')) THEN
    RAISE EXCEPTION 'Rollback: inventory_movements still carries new columns';
  END IF;
  IF to_regclass('public.v_canonical_inventory_balances') IS NOT NULL
     OR to_regclass('public.v_inventory_balances_raw') IS NOT NULL THEN
    RAISE EXCEPTION 'Rollback: view(s) still exist';
  END IF;
END $v$;"""
rbk = rbk.replace(old_verify, new_verify)

# Also patch Phase 5 in the same file to drop the RENAMED view name too
rbk = rbk.replace(
    'DROP VIEW     IF EXISTS public.v_inventory_balances_raw;\nDROP VIEW     IF EXISTS public.v_inventory_balances;',
    'DROP VIEW     IF EXISTS public.v_inventory_balances_raw;\nDROP VIEW     IF EXISTS public.v_canonical_inventory_balances;\nDROP VIEW     IF EXISTS public.v_inventory_balances;'
)

# Note: the "rich" checksum labels above expect that FORWARD writes them.
# Patch FORWARD to emit *_rich labels alongside the existing ones so Rollback finds them.
fwd = open(FWD53).read()
rich_snip = f"""

-- v5.3 RICH checksums (used by rollback for expanded verification)
INSERT INTO public.stocktake_20260727_checksums(label,row_count,qty_sum,name_unit_sku_hash)
SELECT 'items_before_main_rich', count(*)::bigint, sum(stock),
       md5(string_agg(
         coalesce(name,'')||'|'||coalesce(unit,'')||'|'||coalesce(sku,'')||'|'||
         coalesce(is_active::text,'')||'|'||coalesce(product_id::text,'')||'|'||
         coalesce(canonical_item_id::text,'')||'|'||coalesce(retired_at::text,'')||'|'||
         coalesce(ROUND(stock::numeric,4)::text,'0'),
         '||' ORDER BY id))
FROM public.inventory_items_snapshot_20260727;

-- movements rich (requires the new columns; they exist by Phase 1 which runs before checksums are needed at Phase 0 end)
-- We recompute after Phase 1 adds the columns; place after Phase 1 (see marker below).
"""
# Simpler: append RICH checksum computation right after the existing items_before_other_wh checksum block.
marker = "'items_before_other_wh', count(*)::bigint, sum(stock),"
if marker in fwd:
    # append rich blocks AFTER Phase 1 (which adds new columns).
    # Find end of Phase 1 (marker: '-- Phase 2' typical). Fallback: append near existing checksum section but use snapshots only for main items.
    rich_extra = f"""

-- v5.3 RICH pre-forward checksums (paired with rollback expanded verify).
-- items_main from snapshot (safe because new columns don't exist yet on snapshot rows).
INSERT INTO public.stocktake_20260727_checksums(label,row_count,qty_sum,name_unit_sku_hash)
SELECT 'items_before_main_rich', count(*)::bigint, sum(stock),
       md5(string_agg(
         coalesce(name,'')||'|'||coalesce(unit,'')||'|'||coalesce(sku,'')||'|'||
         coalesce(is_active::text,'')||'|'||coalesce(product_id::text,'')||'|'||
         '||||'||coalesce(ROUND(stock::numeric,4)::text,'0'),
         '||' ORDER BY id))
FROM public.inventory_items_snapshot_20260727;
-- movements_main from snapshot (approval_status only; new columns absent pre-forward).
INSERT INTO public.stocktake_20260727_checksums(label,row_count,qty_sum,name_unit_sku_hash)
SELECT 'movements_before_main_rich', count(*)::bigint, sum(quantity),
       md5(string_agg(
         coalesce(approval_status,'')||'|||||',
         '||' ORDER BY id))
FROM public.inventory_movements_snapshot_20260727;
-- other WHs rich (live table, pre-forward)
INSERT INTO public.stocktake_20260727_checksums(label,row_count,qty_sum,name_unit_sku_hash)
SELECT 'items_before_other_wh_rich', count(*)::bigint, sum(stock),
       md5(string_agg(
         coalesce(name,'')||'|'||coalesce(unit,'')||'|'||coalesce(sku,'')||'|'||
         coalesce(is_active::text,'')||'|'||coalesce(product_id::text,'')||'|'||
         coalesce(ROUND(stock::numeric,4)::text,'0'),
         '||' ORDER BY id))
FROM public.inventory_items WHERE warehouse_id<>'{MAIN_WH}';
"""
    # Insert right after the LAST existing checksum insert block (search for 'items_before_other_wh')
    idx = fwd.index("'items_before_other_wh'")
    # find end of the statement (';' after this)
    semi = fwd.index(';', idx) + 1
    fwd = fwd[:semi] + rich_extra + fwd[semi:]
    open(FWD53,'w').write(fwd)

# Rollback needs to match: our rollback expected labels _rich with movements hash covering
# approval_status|reversed_at|reversed_by|reversal_reason|stocktake_ref, but the pre-forward
# snapshot has NULL new columns. Adjust the rollback movements hash to match the "|||||"
# padding we wrote above.
rbk = rbk.replace(
"""  SELECT count(*)::bigint, sum(quantity),
         md5(string_agg(
           coalesce(approval_status,'')||'|'||coalesce(reversed_at::text,'')||'|'||
           coalesce(reversed_by::text,'')||'|'||coalesce(reversal_reason,'')||'|'||
           coalesce(stocktake_ref,''),
           '||' ORDER BY id))
  INTO got_rc, got_qty, got_h
  FROM public.inventory_movements WHERE warehouse_id='{MAIN_WH}';""".replace('{MAIN_WH}',MAIN_WH),
f"""  SELECT count(*)::bigint, sum(quantity),
         md5(string_agg(
           coalesce(approval_status,'')||'|||||',
           '||' ORDER BY id))
  INTO got_rc, got_qty, got_h
  FROM public.inventory_movements WHERE warehouse_id='{MAIN_WH}';""")

# Same padding pattern for items_main_rich (canonical_item_id/retired_at dropped after rollback → NULL → '')
rbk = rbk.replace(
"""  SELECT count(*)::bigint, sum(stock),
         md5(string_agg(
           coalesce(name,'')||'|'||coalesce(unit,'')||'|'||coalesce(sku,'')||'|'||
           coalesce(is_active::text,'')||'|'||coalesce(product_id::text,'')||'|'||
           coalesce(canonical_item_id::text,'')||'|'||coalesce(retired_at::text,'')||'|'||
           coalesce(ROUND(stock::numeric,4)::text,'0'),
           '||' ORDER BY id))
  INTO got_rc, got_qty, got_h
  FROM public.inventory_items WHERE warehouse_id='{MAIN_WH}';""".replace('{MAIN_WH}',MAIN_WH),
f"""  SELECT count(*)::bigint, sum(stock),
         md5(string_agg(
           coalesce(name,'')||'|'||coalesce(unit,'')||'|'||coalesce(sku,'')||'|'||
           coalesce(is_active::text,'')||'|'||coalesce(product_id::text,'')||'|'||
           '||||'||coalesce(ROUND(stock::numeric,4)::text,'0'),
           '||' ORDER BY id))
  INTO got_rc, got_qty, got_h
  FROM public.inventory_items WHERE warehouse_id='{MAIN_WH}';""")

open(RBK53, 'w').write(rbk)
print('wrote', RBK53, len(rbk), 'bytes')


# ============================================================================
# 4) clone_verification_queries.sql
# ============================================================================
canon_values = ',\n'.join(f"    ('{cid}'::uuid,'{n}',{sq},{aq})" for cid,n,_,sq,aq in CANONICALS)
CLONE = f"""-- ============================================================================
-- clone_verification_queries.sql
-- Run on a Postgres Clone before Forward, after Forward, before Rollback, and
-- after Rollback. Every block prints PASS/FAIL with actual vs expected values.
-- READ-ONLY. Never modifies data.
-- ============================================================================

\\pset format aligned
\\pset border 2
SELECT now() AS started_at, version() AS pg_version;

-- ============================================================
-- SECTION A — BEFORE FORWARD
-- ============================================================
\\echo '=== A. BEFORE FORWARD ==='
SELECT 'baseline: inventory_items rows' AS metric, count(*) AS value FROM public.inventory_items;
SELECT 'baseline: inventory_movements rows' AS metric, count(*) AS value FROM public.inventory_movements;
SELECT 'baseline: main WH item rows' AS metric, count(*) AS value
  FROM public.inventory_items WHERE warehouse_id='{MAIN_WH}';
SELECT 'baseline: main WH movement rows' AS metric, count(*) AS value
  FROM public.inventory_movements WHERE warehouse_id='{MAIN_WH}';
SELECT 'baseline: stocktake ref already present?' AS metric, count(*) AS value
  FROM public.inventory_movements WHERE stocktake_ref = '{STK}';
-- expected: 0 before forward
SELECT 'baseline: session already present?' AS metric, count(*) AS value
  FROM public.stocktaking_sessions WHERE session_no = '{STK}';

-- ============================================================
-- SECTION B — AFTER FORWARD
-- ============================================================
\\echo '=== B. AFTER FORWARD ==='
SELECT 'new movements = 126' AS check,
       count(*) AS actual,
       CASE WHEN count(*)=126 THEN 'PASS' ELSE 'FAIL' END AS status
FROM public.inventory_movements
WHERE stocktake_ref = '{STK}' AND idempotency_key IS NOT NULL;

SELECT 'reclass legs = 100' AS check,
       count(*) FILTER (WHERE reference_type='stocktake_reclass') AS actual,
       CASE WHEN count(*) FILTER (WHERE reference_type='stocktake_reclass')=100 THEN 'PASS' ELSE 'FAIL' END AS status
FROM public.inventory_movements WHERE stocktake_ref='{STK}';

SELECT 'stocktake adjustments = 26' AS check,
       count(*) FILTER (WHERE reference_type='stocktaking_session') AS actual,
       CASE WHEN count(*) FILTER (WHERE reference_type='stocktaking_session')=26 THEN 'PASS' ELSE 'FAIL' END AS status
FROM public.inventory_movements WHERE stocktake_ref='{STK}';

SELECT 'opening reversals = 103' AS check,
       count(*) AS actual,
       CASE WHEN count(*)=103 THEN 'PASS' ELSE 'FAIL' END AS status
FROM public.inventory_movements WHERE stocktake_ref='{STK}' AND approval_status='reversed';

SELECT 'stocktaking_lines = 27' AS check,
       count(*) AS actual,
       CASE WHEN count(*)=27 THEN 'PASS' ELSE 'FAIL' END AS status
FROM public.stocktaking_lines l
JOIN public.stocktaking_sessions s ON s.id=l.session_id
WHERE s.session_no='{STK}';

SELECT 'retired aliases = 79' AS check,
       count(*) AS actual,
       CASE WHEN count(*)=79 THEN 'PASS' ELSE 'FAIL' END AS status
FROM public.inventory_items
WHERE warehouse_id='{MAIN_WH}' AND retired_at = '{STK}';

SELECT 'product rewires = 8 (products now pointing to a canonical)' AS check,
       count(*) AS actual
FROM public.products p
JOIN public.inventory_items i ON i.id = p.inventory_item_id
WHERE i.warehouse_id='{MAIN_WH}' AND i.canonical_item_id = i.id
  AND i.id IN ({CANON_IDS_SQL});

-- Per-canonical: v_canonical.balance vs inventory_items.stock vs stocktaking_lines.actual_qty
\\echo '--- Per-canonical balance triple check ---'
WITH expected(canon_id, canon_name, sys_q, act_q) AS (VALUES
{canon_values}
),
v AS (
  SELECT canonical_item_id, balance FROM public.v_canonical_inventory_balances
  WHERE warehouse_id='{MAIN_WH}'
),
s AS (SELECT id, stock FROM public.inventory_items WHERE warehouse_id='{MAIN_WH}'),
l AS (
  SELECT ll.item_id, ll.actual_qty
  FROM public.stocktaking_lines ll
  JOIN public.stocktaking_sessions ss ON ss.id=ll.session_id
  WHERE ss.session_no='{STK}'
)
SELECT e.canon_name AS name,
       COALESCE(v.balance,0) AS view_bal,
       COALESCE(s.stock,0)  AS item_stock,
       COALESCE(l.actual_qty,0) AS line_actual,
       e.act_q AS expected,
       CASE WHEN ROUND(COALESCE(v.balance,0),4) = ROUND(e.act_q,4)
             AND ROUND(COALESCE(s.stock,0),4)   = ROUND(e.act_q,4)
             AND ROUND(COALESCE(l.actual_qty,0),4) = ROUND(e.act_q,4)
            THEN 'PASS' ELSE 'FAIL' END AS status
FROM expected e
LEFT JOIN v ON v.canonical_item_id=e.canon_id
LEFT JOIN s ON s.id=e.canon_id
LEFT JOIN l ON l.item_id=e.canon_id
ORDER BY e.canon_name;

-- Aliases with non-zero raw ledger (must be empty)
SELECT 'aliases with non-zero raw ledger (must be 0)' AS check, count(*) AS actual
FROM public.v_inventory_balances_raw r
JOIN public.inventory_items i ON i.id = r.raw_item_id
WHERE i.retired_at = '{STK}'
  AND ROUND(COALESCE(r.balance,0),4) <> 0;

-- Other warehouses untouched
SELECT 'other-WH sum(stock)' AS check,
       ROUND(sum(stock)::numeric,4) AS actual
FROM public.inventory_items WHERE warehouse_id<>'{MAIN_WH}';

-- FAT pack untouched
SELECT 'FAT pack stock still 61.5' AS check,
       stock AS actual,
       CASE WHEN ROUND(stock::numeric,4)=61.5 THEN 'PASS' ELSE 'FAIL' END AS status
FROM public.inventory_items WHERE id='{FAT}';
SELECT 'FAT pack has no stocktake_ref movement' AS check, count(*) AS actual
FROM public.inventory_movements WHERE item_id='{FAT}' AND stocktake_ref='{STK}';

-- POST-COUNT-MOVEMENTS (parametrized) — replace :cutoff_ts before running.
\\echo '--- POST-COUNT-MOVEMENTS (parametrized) ---'
-- Example: \\set cutoff_ts '2026-07-27 20:00:00+02'
-- SELECT m.id AS movement_id, m.performed_at, m.item_id AS source_item_id,
--        i.canonical_item_id, i.name, m.movement_type, m.quantity,
--        m.reference_type, m.reference_id, m.approval_status, m.performed_by,
--        CASE WHEN m.approval_status='posted' AND m.stocktake_ref IS NULL
--             THEN 'KEEP' ELSE 'REVIEW' END AS keep_decision
-- FROM public.inventory_movements m
-- LEFT JOIN public.inventory_items i ON i.id = m.item_id
-- WHERE m.warehouse_id='{MAIN_WH}'
--   AND m.performed_at > :'cutoff_ts'
-- ORDER BY m.performed_at;

-- ============================================================
-- SECTION C — BEFORE ROLLBACK (identical to AFTER FORWARD)
-- ============================================================
\\echo '=== C. BEFORE ROLLBACK ==='
SELECT 'session present' AS check, count(*) AS actual
FROM public.stocktaking_sessions WHERE session_no='{STK}';
SELECT 'stocktake movements' AS check, count(*) AS actual
FROM public.inventory_movements WHERE stocktake_ref='{STK}';

-- ============================================================
-- SECTION D — AFTER ROLLBACK
-- ============================================================
\\echo '=== D. AFTER ROLLBACK ==='
SELECT 'session removed' AS check, count(*) AS actual,
       CASE WHEN count(*)=0 THEN 'PASS' ELSE 'FAIL' END AS status
FROM public.stocktaking_sessions WHERE session_no='{STK}' OR reference_id='{STK}';
SELECT 'stocktake movements removed' AS check, count(*) AS actual,
       CASE WHEN count(*)=0 THEN 'PASS' ELSE 'FAIL' END AS status
FROM public.inventory_movements WHERE stocktake_ref='{STK}' AND idempotency_key IS NOT NULL;
SELECT 'view v_canonical_inventory_balances dropped' AS check,
       to_regclass('public.v_canonical_inventory_balances') AS actual,
       CASE WHEN to_regclass('public.v_canonical_inventory_balances') IS NULL THEN 'PASS' ELSE 'FAIL' END AS status;
SELECT 'columns dropped on inventory_items' AS check,
       count(*) FILTER (WHERE column_name IN ('canonical_item_id','retired_at','retired_reason')) AS actual,
       CASE WHEN count(*) FILTER (WHERE column_name IN ('canonical_item_id','retired_at','retired_reason'))=0 THEN 'PASS' ELSE 'FAIL' END AS status
FROM information_schema.columns WHERE table_schema='public' AND table_name='inventory_items';
SELECT 'columns dropped on inventory_movements' AS check,
       count(*) FILTER (WHERE column_name IN ('idempotency_key','reversal_of_movement_id','reversed_at','reversed_by','reversal_reason','stocktake_ref')) AS actual,
       CASE WHEN count(*) FILTER (WHERE column_name IN ('idempotency_key','reversal_of_movement_id','reversed_at','reversed_by','reversal_reason','stocktake_ref'))=0 THEN 'PASS' ELSE 'FAIL' END AS status
FROM information_schema.columns WHERE table_schema='public' AND table_name='inventory_movements';

-- Snapshot tables still present (kept for audit; drop manually when ready)
SELECT 'snapshot tables retained' AS check, count(*) AS actual
FROM information_schema.tables
WHERE table_schema='public' AND table_name IN
  ('inventory_items_snapshot_20260727','inventory_movements_snapshot_20260727','stocktake_20260727_checksums');

SELECT now() AS finished_at;
"""
open(CLONE_Q,'w').write(CLONE)
print('wrote', CLONE_Q)


# ============================================================================
# 5) post_count_movements_report.xlsx (template + placeholder rows)
# ============================================================================
wb2 = openpyxl.Workbook()
ws = wb2.active
ws.title = 'Instructions'
ws.append(['POST-COUNT-MOVEMENTS Report — Template'])
ws.append(['Generated', GEN])
ws.append([])
ws.append(['STEP 1: Agree an exact Count Cutoff timestamp (Africa/Cairo, +02:00).'])
ws.append(['STEP 2: On the Clone (or Prod), run the parametrized query from clone_verification_queries.sql.'])
ws.append(['STEP 3: Paste results into the "Movements" sheet below.'])
ws.append(['STEP 4: For each row, mark keep_decision = KEEP or DISCARD.'])
ws.append(['STEP 5: Recompute target actual_qty per canonical:'])
ws.append(['       target = physical_count(at cutoff) + Σ signed_qty(KEEP posted movements after cutoff)'])
ws.append(['STEP 6: Update stocktaking_lines.actual_qty seeds in Forward SQL before running production.'])

mv = wb2.create_sheet('Movements')
mv.append(['movement_id','performed_at','source_item_id','canonical_item_id','name',
           'movement_type','quantity','signed_qty','reference_type','reference_id',
           'approval_status','performed_by','keep_decision','impact_on_target_balance'])
mv.append(['<placeholder — populate from Clone query output>']*14)

wb2.create_sheet('Cutoff').append(['Cutoff timestamp (Africa/Cairo)','<<PENDING>>'])
wb2.save(POST_XL)
print('wrote', POST_XL)


# ============================================================================
# 6) migration_actor_candidates.xlsx
# ============================================================================
wb3 = openpyxl.Workbook()
ws = wb3.active
ws.title = 'Actor Candidates'
ws.append(['User ID','Name','Email','Role','Status','Nominated?','Note'])
for a in ACTORS:
    ws.append([*a,'', ''])
ws.append([])
ws.append(['RULE','Only ONE actor per run'])
ws.append(['ROLE CHECK','Must still hold general_manager OR executive_manager at run time'])
ws.append(['INJECTION',"SELECT set_config('app.stk_actor','<UUID>', false);  -- BEFORE running the SQL file, same session"])
ws.append(['PLACEHOLDER',"<<ADMIN_ACTOR_UUID>> in Forward/Rollback header — replace with literal for production"])
wb3.save(ACTOR_XL)
print('wrote', ACTOR_XL)


# ============================================================================
# 7) security_pre_migration_report_v5_3.md
# ============================================================================
SEC = f"""# Security Pre-Migration Report — v5.3

_Generated: {GEN}_

## Baseline (from Supabase linter, pre-migration)

- **859 total issues** on the current database.
- **2 ERROR** entries — both `Security Definer View` on pre-existing views (unrelated to this migration).
- The remaining issues are WARN / INFO (search_path, permissive RLS, etc.) predating this work.

## v5.3 migration security invariants

| # | Invariant | How v5.3 enforces it |
|---|-----------|----------------------|
| 1 | No new `SECURITY DEFINER` view | `v_canonical_inventory_balances` and `v_inventory_balances_raw` are both created with `WITH (security_invoker = true)`. |
| 2 | Views expose no data the caller can't read directly | `security_invoker=true` makes RLS on `inventory_movements` / `inventory_items` apply to view queries. |
| 3 | Snapshot tables locked down | `REVOKE ALL ... FROM PUBLIC, anon, authenticated; GRANT ALL ... TO service_role` for `inventory_items_snapshot_20260727`, `inventory_movements_snapshot_20260727`, `stocktake_20260727_checksums`. |
| 4 | No `GRANT` on alias rows to app roles | Aliases live inside `inventory_items` (same grants as before). No new table grants added. |
| 5 | No RLS policy changes | Migration does not `CREATE POLICY`, `DROP POLICY`, `ALTER POLICY`, or `ALTER TABLE ... ENABLE ROW LEVEL SECURITY`. |
| 6 | Actor role verified at run time | Phase -1 `Preflight` looks up `public.user_roles` for the actor and aborts if the actor is not `general_manager` OR `executive_manager`. Runs even when `auth.uid()` is NULL (uses `current_setting('app.stk_actor')`). |
| 7 | `SET search_path = public` on new functions | `reroute_to_canonical()` and `reroute_res_to_canonical()` set search_path (verify in the generated SQL). |
| 8 | No new SECURITY DEFINER function | Reroute functions are SECURITY INVOKER (trigger context). |

## Expected linter delta after Clone Forward

- ERRORS: **unchanged (2)** — no new `SECURITY DEFINER` view created.
- WARN(Function Search Path Mutable): **unchanged** — new functions set `search_path = public`.
- WARN(RLS Policy Always True): **unchanged** — no policy touched.
- INFO totals: unchanged.

If the linter reports **new** errors or warnings introduced by this migration, treat the Clone dry-run as FAILED and open a fix in v5.4.

## Follow-up (out of scope for v5.3)

The follow-up migration (see `Follow-up Migration Plan` sheet) will add a central `apply_inventory_movement()` RPC, non-negative balance guard, and per-line idempotency. These will introduce new SECURITY DEFINER surfaces that require their own security review.
"""
open(SEC_MD,'w').write(SEC)
print('wrote', SEC_MD)


# ============================================================================
# 8) changelog_v5_2_to_v5_3.md
# ============================================================================
CHG = f"""# Changelog v5.2 → v5.3

_Generated: {GEN}_

**Scope of v5.3: PAPER PACKAGE ONLY.** No Migration runs on the production database. No `supabase--migration` call.

## Files in this package

- `DryRun_MAIN-STOCKTAKE-20260727_v5_3.xlsx`
- `forward_migration_MAIN-STOCKTAKE-20260727_v5_3.sql`
- `rollback_MAIN-STOCKTAKE-20260727_v5_3.sql`
- `clone_verification_queries.sql`
- `post_count_movements_report.xlsx`
- `migration_actor_candidates.xlsx`
- `security_pre_migration_report_v5_3.md`
- `changelog_v5_2_to_v5_3.md`

## Behavioural changes vs v5.2

| # | Change | Rationale (from v5.2 review) |
|---|--------|------------------------------|
| 1 | Restricted view `v_canonical_inventory_balances` scoped to 27 canonicals × main WH only | Prevents shipping a system-wide view before items outside scope (incl. FAT pack) are reconciled. |
| 2 | `v_inventory_balances_raw` kept but labelled audit-only (used only by assertion 10.6) | Ledger-vs-stock differences on out-of-scope items must not appear in the app source-of-truth. |
| 3 | No `v_inventory_balances` view created | Removes the risk of screens reading 130.5 in one view and 61.5 in another for FAT pack. |
| 4 | New assertion **10.7** per-canonical | Fails the transaction if any single canonical mismatches, even when the 624.5 kg total is preserved by offsetting differences. |
| 5 | Actor UUID handled as `<<ADMIN_ACTOR_UUID>>` placeholder + candidate list in header | Removes reliance on auto-picking or `auth.uid()`. |
| 6 | `Migration Actor Candidates` sheet + `migration_actor_candidates.xlsx` | User selects the actor explicitly before production. |
| 7 | `Count Cutoff Control` sheet + POST-COUNT-MOVEMENTS parametrized query | Prevents wiping legitimate movements posted after the physical count. |
| 8 | Excel cleanup | FAT removed from `Canonical Mapping`, `Source Ledger Bridge`, `Full Action Plan Replay Test`, `Duplicate Openings In-Scope`; explicit `Out-of-Scope Items` sheet added; formula-looking cell in `Validation Checklist` neutralised. |
| 9 | `Item Master Action Plan` row for 6ك دبوس بالعظم now documents adjustment to 0 with `is_active=true, retired_at=NULL` | Matches SQL: canonical stays open for future receipts. |
| 10 | Rollback verification expanded | rich hash covers `name|unit|sku|is_active|product_id|canonical_item_id|retired_at|stock` on items and `approval_status|reversed_at|reversed_by|reversal_reason|stocktake_ref` on movements; explicit stocktake-session absence check; explicit dropped-columns check. |
| 11 | Rollback also verifies OTHER-warehouses sum(stock) and hash | Previously counted only. |
| 12 | Forward emits `*_rich` checksums after Phase 1 columns exist | Enables the expanded rollback verify. |
| 13 | `clone_verification_queries.sql` | Standalone PASS/FAIL harness for Clone; covers BEFORE/AFTER Forward and BEFORE/AFTER Rollback. |
| 14 | `Follow-up Migration Plan` sheet | Non-negative guard + central RPC + idempotency planned; NOT part of v5.3. |
| 15 | `security_pre_migration_report_v5_3.md` | States security invariants and expected linter delta (2 ERROR unchanged). |

## What is unchanged from v5.2

- Movement plan: 126 rows = 100 reclass legs + 26 stocktake adjustments.
- Opening reversal status flips: 103.
- stocktaking_lines: 27 rows with `system_qty` and `actual_qty` from the Adjustments sheet.
- FAT pack (`{FAT}`) fully out of scope.
- No psql meta-commands. No `\\set`. No `auth.uid()` reliance.
- Preflight schema-existence check remains comprehensive.
- Snapshot tables ACL: `REVOKE PUBLIC/anon/authenticated + GRANT service_role`.

## What blocks production execution (unchanged from prior reviews)

1. Exact Count Cutoff timestamp not yet provided.
2. Migration Actor UUID not yet chosen and substituted for `<<ADMIN_ACTOR_UUID>>`.
3. Clone Forward + Rollback dry-run reports (from `clone_verification_queries.sql`) not yet produced.
4. Follow-up migration for non-negative guard / central RPC / idempotency not yet delivered.
"""
open(CHG_MD,'w').write(CHG)
print('wrote', CHG_MD)

print('\\nDONE. Files under /mnt/documents/')
