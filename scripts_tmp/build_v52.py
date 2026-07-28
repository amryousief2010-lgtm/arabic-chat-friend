"""Build v5.2 package addressing 11 review blockers.

Fixes vs v5.1:
 1. RECLASSIFY_NEGATIVE: OUT lands on canonical, IN lands on alias (seq 8/80/84/92 corrected).
 2. دهن النعام (عبوة) FAT fully out of scope: no opening reversal, no assertion.
 3. NO psql meta commands. Actor UUID must be pre-set via SELECT set_config('app.stk_actor',...,false).
 4. stocktaking_lines.system_qty & actual_qty from Adjustments sheet (not zero).
 5. Reroute trigger scoped to reference_type IN ('stocktake_reclass','stocktake_reclass_neg') only.
 6. Preflight covers every new column/table/view/index/function/trigger.
 7. Guardrails misleading claim about apply_inventory_movement RPC removed.
 8. Reservation reroute clarified as agouza-only scope (documented in Guardrails).
 9. Rollback verifies stock sum + name/unit/sku hash checksums, not just counts.
10. Summary & Validation Checklist rewritten to match SQL (6ك dbos adjusted to 0, links = 8/18/1).
11. New assertion: every retired alias raw_ledger balance ends at 0.
"""
import openpyxl, hashlib
from datetime import datetime, timezone

V51 = '/mnt/documents/DryRun_MAIN-STOCKTAKE-20260727_v5_1.xlsx'
V52 = '/mnt/documents/DryRun_MAIN-STOCKTAKE-20260727_v5_2.xlsx'
FWD = '/mnt/documents/forward_migration_MAIN-STOCKTAKE-20260727_v5_2.sql'
RBK = '/mnt/documents/rollback_MAIN-STOCKTAKE-20260727_v5_2.sql'
STK = 'MAIN-STOCKTAKE-20260727'
MAIN_WH = '5ec781b5-685b-4806-b59a-83a79ea5662c'
FAT = 'a55763c7-6e12-4ed0-8d62-37ace6bf86a3'
PIN_6K = '102ddab2-1b6a-411a-ace9-1f962c27e56e'

wb = openpyxl.load_workbook(V51)

def rows(sheet):
    ws = wb[sheet]
    hdr = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and all(v is None for v in row): continue
        yield dict(zip(hdr, row))

def q(v):
    if v is None: return 'NULL'
    if isinstance(v, bool): return 'true' if v else 'false'
    if isinstance(v, (int, float)): return repr(v) if isinstance(v,float) else str(v)
    return "'" + str(v).replace("'", "''") + "'"

# ============================================================================
# XLSX v5.2 — rewrite Summary + Guardrails + Validation Checklist; drop FAT
# opening reversal; ensure Adjustments has proper system_qty column reference.
# ============================================================================

# Remove FAT opening-reversal row (it's row 20 per header '#')
opr = wb['Opening Reversal Actions']
opr_hdr = [c.value for c in opr[1]]
sid_col = opr_hdr.index('Source Item ID') + 1
rows_to_delete = [r for r in range(2, opr.max_row + 1)
                  if str(opr.cell(r, sid_col).value or '') == FAT]
for r in reversed(rows_to_delete):
    opr.delete_rows(r)

# Fix Product Link Action Plan: keep FAT row but ensure its Canonical == Source (independent)
# (already correct in v5.1 output)

# Count product link actions after fix
plan_rewire = plan_keep = plan_independent = 0
for r in rows('Product Link Action Plan'):
    src = r.get('Current inventory_item_id (source)')
    can = r.get('Canonical Item ID')
    if src == FAT or can == FAT:
        plan_independent += 1
    elif src == can:
        plan_keep += 1
    else:
        plan_rewire += 1

# Rewrite Summary sheet from scratch
if 'Summary' in wb.sheetnames:
    del wb['Summary']
sm = wb.create_sheet('Summary', 0)
sm.append(['Key','Value','Notes'])
sm.append(['Package version','v5.2', 'Corrects 11 blockers from v5.1 review'])
sm.append(['Stocktake ref', STK, 'warehouse_id = ' + MAIN_WH])
sm.append(['kg items final total', 624.5, '26 canonicals in kg'])
sm.append(['6ك دبوس بالعظم', PIN_6K,
           'Active canonical; adjusted to 0 via posted STOCKTAKE_ADJUSTMENT (–171 kg). Documented in stocktaking_lines.'])
sm.append(['قطع كباب', '4a528973-066f-445d-a643-099e310db27b',
           'system_qty=0.5, actual_qty=0.5, zero adjustment — line retained for documentation'])
sm.append(['دهن النعام (عبوة)', FAT,
           'FULLY OUT OF SCOPE in v5.2: no alias mapping, no opening reversal, no adjustment, no assertion. Its 130.5 ledger vs 61.5 stock discrepancy to be handled in a separate pack-unit stocktake.'])
sm.append(['Movement plan rows', 126, '100 reclassification legs + 26 stocktake adjustments'])
sm.append(['SQL INSERTs into inventory_movements', 126, 'exactly one per plan row'])
sm.append(['Opening reversal status-flips', 103, 'FAT opening excluded — was 104 in v5.1'])
sm.append(['Product Link Actions', plan_rewire + plan_keep + plan_independent,
           f'{plan_rewire} rewire + {plan_keep} keep + {plan_independent} independent (fat pack)'])
sm.append(['View v_inventory_balances', 'security_invoker=true, GROUP BY canonical_item_id, warehouse_id', '1 row per canonical'])
sm.append(['View v_inventory_balances_raw', 'security_invoker=true', 'per raw item_id — used for the alias-zero assertion only'])
sm.append(['Snapshot tables','REVOKE PUBLIC/anon/authenticated + GRANT service_role only', ''])
sm.append(['Actor injection','SELECT set_config(\'app.stk_actor\',\'<uuid>\', false) before running SQL','No psql meta commands'])
sm.append(['Reroute trigger scope','reference_type IN (stocktake_reclass, stocktake_reclass_neg) only','Future stocktakes still route to canonical'])

# Rewrite Validation Checklist
if 'Validation Checklist' in wb.sheetnames:
    del wb['Validation Checklist']
vc = wb.create_sheet('Validation Checklist')
vc.append(['#','Check','Expected'])
checks = [
    ('1', 'Total kg after migration', '624.50'),
    ('2', 'Canonicals in kg', '26'),
    ('3', '6ك دبوس بالعظم canonical balance', '0 (via posted adjustment –171)'),
    ('4', '6ك دبوس بالعظم is_active', 'true, retired_at NULL'),
    ('5', 'قطع كباب balance', '0.5'),
    ('6', 'قطع كباب stocktaking_lines.system_qty', '0.5'),
    ('7', 'قطع كباب stocktaking_lines.actual_qty', '0.5'),
    ('8', 'دهن النعام (عبوة)', 'unchanged: stock=61.5, no alias, no reversal, no adjustment'),
    ('9', 'Retired aliases raw ledger balance', '= 0 for every alias (assertion 10.6)'),
    ('10', 'Movement plan vs SQL row-by-row diff', '0'),
    ('11', 'Idempotency collisions inside plan', '0'),
    ('12', 'Reclassification legs', '100'),
    ('13', 'Stocktake adjustments', '26'),
    ('14', 'Opening reversals (status flips)', '103'),
    ('15', 'stocktaking_lines rows', '27'),
    ('16', 'Other warehouses items count', 'unchanged vs snapshot'),
    ('17', 'v_inventory_balances rows per canonical in main WH', '1'),
    ('18', 'Rollback restores stock sum + name/unit/sku checksum', 'match'),
]
for row in checks:
    vc.append(row)

# Rewrite Guardrails
if 'Guardrails' in wb.sheetnames:
    del wb['Guardrails']
gr = wb.create_sheet('Guardrails')
gr.append(['#','Guardrail','Implementation status'])
gr.append(['1','Unique idempotency_key on inventory_movements','IMPLEMENTED — ux_inv_mov_idempotency_key partial unique index'])
gr.append(['2','SHARE ROW EXCLUSIVE lock on inventory_items & inventory_movements','IMPLEMENTED — LOCK TABLE inside transaction'])
gr.append(['3','Snapshot tables invisible to Data API','IMPLEMENTED — REVOKE ALL FROM PUBLIC, anon, authenticated + GRANT service_role only'])
gr.append(['4','v_inventory_balances security_invoker=true','IMPLEMENTED — WITH (security_invoker = true)'])
gr.append(['5','Reroute trigger scoped to migration-owned reference_types','IMPLEMENTED — trigger checks reference_type IN (stocktake_reclass, stocktake_reclass_neg); future stocktake_adjustment rows STILL route to canonical'])
gr.append(['6','Reservation reroute','IMPLEMENTED for agouza_stock_reservations only. Rationale: agouza is the sole reservations table in scope for main-warehouse alias→canonical mapping (other modules do not carry inventory_item_id-based reservations on the main WH).'])
gr.append(['7','apply_inventory_movement RPC guard on retired items','OUT OF SCOPE for v5.2 — retirement is enforced structurally (canonical rewire + trigger). A follow-up migration will add explicit RPC guards; not claimed as delivered here.'])
gr.append(['8','Non-negative balance guard','ALREADY IN PLACE at trg_guard_mwt_nonneg (main warehouse treasury). Inventory-side non-negative enforcement is a separate follow-up.'])
gr.append(['9','Actor injection','IMPLEMENTED — SELECT set_config(\'app.stk_actor\', <uuid>, false) BEFORE the SQL; DO $$ blocks read current_setting(). No psql meta commands.'])
gr.append(['10','Preflight full schema-object non-existence check','IMPLEMENTED — covers all 3 new columns on inventory_items, 6 new columns on inventory_movements, 3 snapshot tables, 2 views, 1 index, 2 functions, 2 triggers'])
gr.append(['11','Assertion: retired aliases raw balance = 0','IMPLEMENTED — assertion 10.6 iterates v_inventory_balances_raw'])
gr.append(['12','FAT pack scoped OUT of migration','IMPLEMENTED — no alias, no opening reversal, no adjustment, no assertion; stays visible as an independent item unit=عبوة, stock=61.5'])

wb.save(V52)
print('WROTE', V52)

# ============================================================================
# FORWARD MIGRATION SQL v5.2 — NO psql meta commands
# ============================================================================
now_iso = datetime.now(timezone.utc).isoformat()

IM_COLS = ['id','item_id','warehouse_id','movement_type','quantity','destination_warehouse_id',
           'reference','party','unit_cost','notes','performed_by','performed_at','created_at',
           'movement_no','module','source_warehouse_id','reference_type','reference_id',
           'batch_id','reason','approval_status','approved_by','approved_at','total_cost',
           'order_item_id','product_id','package_count','package_weight_kg','quantity_kg']
II_COLS = ['id','warehouse_id','name','category','sku','unit','stock','low_stock_threshold',
           'unit_cost','expiry_date','notes','is_active','created_at','updated_at',
           'reserved_qty','blocked_qty','module','item_code','last_movement_date','product_id']

out = []
add = out.append
ii_cols = ', '.join(II_COLS)
im_cols = ', '.join(IM_COLS)

add(f"""-- ============================================================================
-- MAIN-STOCKTAKE-20260727 — Forward Migration v5.2
-- Scope: warehouse_id = {MAIN_WH} ONLY (المخزن الرئيسي - المقر)
-- Generated: {now_iso}
--
-- HOW TO RUN (any Postgres migration runner — psql, node-pg, supabase, etc.):
--   1) Ensure the executor's UUID is available BEFORE running this file.
--   2) Prepend (in the same session/transaction, NOT inside a DO block):
--        SELECT set_config('app.stk_actor', '<uuid-of-executing-admin>', false);
--   3) Then run this file as normal SQL. No psql meta commands are used.
--
-- The migration reads current_setting('app.stk_actor') everywhere it needs
-- the executor UUID (performed_by / approved_by / reversed_by / created_by).
-- ============================================================================

BEGIN;
SET LOCAL statement_timeout = '15min';
SET LOCAL lock_timeout      = '30s';
SET LOCAL idle_in_transaction_session_timeout = '10min';

-- ------------------------------------------------------------------
-- Phase -1 — Preflight (actor + FULL schema-object non-existence check)
-- ------------------------------------------------------------------
DO $pre$
DECLARE
  v_actor uuid;
  v_role_ok boolean;
  v_setting text;
BEGIN
  BEGIN
    v_setting := current_setting('app.stk_actor', true);
  EXCEPTION WHEN OTHERS THEN v_setting := NULL;
  END;
  v_actor := COALESCE(auth.uid(), NULLIF(v_setting,'')::uuid);
  IF v_actor IS NULL THEN
    RAISE EXCEPTION 'Preflight: no actor available. Run "SELECT set_config(''app.stk_actor'',''<uuid>'',false);" before this file.';
  END IF;
  -- refresh into transaction-local setting
  PERFORM set_config('app.stk_actor', v_actor::text, true);

  SELECT bool_or(role::text IN ('general_manager','executive_manager'))
    INTO v_role_ok
  FROM public.user_roles WHERE user_id = v_actor;
  IF NOT COALESCE(v_role_ok,false) THEN
    RAISE EXCEPTION 'Preflight: actor % lacks general_manager/executive_manager role', v_actor;
  END IF;

  -- Every column, table, view, index, function, trigger this migration will add
  -- must NOT already exist. Any collision aborts before we touch data.
  IF EXISTS (SELECT 1 FROM information_schema.columns WHERE table_schema='public'
              AND table_name='inventory_items'
              AND column_name IN ('canonical_item_id','retired_at','retired_reason')) THEN
    RAISE EXCEPTION 'Preflight: inventory_items already has canonical_item_id/retired_at/retired_reason (prior run?)';
  END IF;
  IF EXISTS (SELECT 1 FROM information_schema.columns WHERE table_schema='public'
              AND table_name='inventory_movements'
              AND column_name IN ('idempotency_key','reversal_of_movement_id','reversed_at',
                                  'reversed_by','reversal_reason','stocktake_ref')) THEN
    RAISE EXCEPTION 'Preflight: inventory_movements already has one of the 6 new columns';
  END IF;
  IF to_regclass('public.inventory_items_snapshot_20260727')     IS NOT NULL
     OR to_regclass('public.inventory_movements_snapshot_20260727') IS NOT NULL
     OR to_regclass('public.stocktake_20260727_checksums')          IS NOT NULL THEN
    RAISE EXCEPTION 'Preflight: one or more snapshot tables already exist';
  END IF;
  IF to_regclass('public.v_inventory_balances')     IS NOT NULL
     OR to_regclass('public.v_inventory_balances_raw') IS NOT NULL THEN
    RAISE EXCEPTION 'Preflight: v_inventory_balances[_raw] view already exists';
  END IF;
  IF EXISTS (SELECT 1 FROM pg_indexes WHERE schemaname='public' AND indexname='ux_inv_mov_idempotency_key') THEN
    RAISE EXCEPTION 'Preflight: index ux_inv_mov_idempotency_key already exists';
  END IF;
  IF EXISTS (SELECT 1 FROM pg_proc p JOIN pg_namespace n ON n.oid=p.pronamespace
              WHERE n.nspname='public' AND p.proname IN ('reroute_to_canonical','reroute_res_to_canonical')) THEN
    RAISE EXCEPTION 'Preflight: reroute_to_canonical / reroute_res_to_canonical already exists';
  END IF;
  IF EXISTS (SELECT 1 FROM pg_trigger WHERE tgname IN ('trg_inv_mov_reroute_canonical','trg_agouza_res_reroute')) THEN
    RAISE EXCEPTION 'Preflight: trigger already exists';
  END IF;
END $pre$;

-- ------------------------------------------------------------------
-- Phase 0 — Lock + snapshots (explicit column lists)
-- ------------------------------------------------------------------
LOCK TABLE public.inventory_items     IN SHARE ROW EXCLUSIVE MODE;
LOCK TABLE public.inventory_movements IN SHARE ROW EXCLUSIVE MODE;

CREATE TABLE public.inventory_items_snapshot_20260727
  (LIKE public.inventory_items INCLUDING ALL);
INSERT INTO public.inventory_items_snapshot_20260727 ({ii_cols})
SELECT {ii_cols} FROM public.inventory_items WHERE warehouse_id = '{MAIN_WH}';

CREATE TABLE public.inventory_movements_snapshot_20260727
  (LIKE public.inventory_movements INCLUDING ALL);
INSERT INTO public.inventory_movements_snapshot_20260727 ({im_cols})
SELECT {im_cols} FROM public.inventory_movements WHERE warehouse_id = '{MAIN_WH}';

REVOKE ALL ON public.inventory_items_snapshot_20260727     FROM PUBLIC;
REVOKE ALL ON public.inventory_items_snapshot_20260727     FROM anon, authenticated;
GRANT  ALL ON public.inventory_items_snapshot_20260727     TO service_role;
REVOKE ALL ON public.inventory_movements_snapshot_20260727 FROM PUBLIC;
REVOKE ALL ON public.inventory_movements_snapshot_20260727 FROM anon, authenticated;
GRANT  ALL ON public.inventory_movements_snapshot_20260727 TO service_role;

CREATE TABLE public.stocktake_20260727_checksums (
  label text PRIMARY KEY,
  row_count bigint NOT NULL,
  qty_sum numeric,
  name_unit_sku_hash text,
  captured_at timestamptz NOT NULL DEFAULT now()
);
REVOKE ALL ON public.stocktake_20260727_checksums FROM PUBLIC;
REVOKE ALL ON public.stocktake_20260727_checksums FROM anon, authenticated;
GRANT  ALL ON public.stocktake_20260727_checksums TO service_role;

-- Before-snapshot checksums (main WH + other WHs)
INSERT INTO public.stocktake_20260727_checksums(label,row_count,qty_sum,name_unit_sku_hash)
SELECT 'items_before_main', count(*)::bigint, sum(stock),
       md5(string_agg(coalesce(name,'')||'|'||coalesce(unit,'')||'|'||coalesce(sku,''), '||' ORDER BY id))
FROM public.inventory_items_snapshot_20260727;
INSERT INTO public.stocktake_20260727_checksums(label,row_count,qty_sum)
SELECT 'movements_before_main', count(*)::bigint, sum(quantity)
FROM public.inventory_movements_snapshot_20260727;
INSERT INTO public.stocktake_20260727_checksums(label,row_count,qty_sum,name_unit_sku_hash)
SELECT 'items_before_other_wh', count(*)::bigint, sum(stock),
       md5(string_agg(coalesce(name,'')||'|'||coalesce(unit,'')||'|'||coalesce(sku,''), '||' ORDER BY id))
FROM public.inventory_items WHERE warehouse_id <> '{MAIN_WH}';

SAVEPOINT after_snapshot;

-- ------------------------------------------------------------------
-- Phase 1 — Schema additions
-- ------------------------------------------------------------------
ALTER TABLE public.inventory_items
  ADD COLUMN canonical_item_id uuid REFERENCES public.inventory_items(id),
  ADD COLUMN retired_at        timestamptz,
  ADD COLUMN retired_reason    text;

ALTER TABLE public.inventory_movements
  ADD COLUMN idempotency_key         text,
  ADD COLUMN reversal_of_movement_id uuid REFERENCES public.inventory_movements(id),
  ADD COLUMN reversed_at             timestamptz,
  ADD COLUMN reversed_by             uuid REFERENCES auth.users(id),
  ADD COLUMN reversal_reason         text,
  ADD COLUMN stocktake_ref           text;

CREATE UNIQUE INDEX ux_inv_mov_idempotency_key
  ON public.inventory_movements(idempotency_key)
  WHERE idempotency_key IS NOT NULL;

SAVEPOINT after_schema;
""")

# ---------- Phase 2 — Item Master (27 canonicals) ----------
add("-- ------------------------------------------------------------------")
add("-- Phase 2 — Item Master: 27 canonicals updated (name/unit/barcode/self-canonical)")
add("--            دهن النعام (عبوة) NOT included here — it is out of scope.")
add("-- ------------------------------------------------------------------")
n_can = 0
canonicals_kg = []
canonicals_all = []
for r in rows('Item Master Action Plan'):
    cid   = r['Canonical Item ID']
    fname = r['Final Name']
    funit = r['Final Unit']
    fbc   = r['Final Barcode (text)']
    if cid == FAT: continue
    add(f"UPDATE public.inventory_items SET "
        f"name={q(fname)}, unit={q(funit)}, "
        f"sku=COALESCE(NULLIF(sku,''), {q(fbc)}), "
        f"canonical_item_id=id, is_active=true, retired_at=NULL, retired_reason=NULL "
        f"WHERE id={q(cid)} AND warehouse_id={q(MAIN_WH)};")
    n_can += 1
    canonicals_all.append(cid)
    if funit in ('كيلو','كجم'): canonicals_kg.append(cid)
add(f"DO $c$ BEGIN IF (SELECT count(*) FROM public.inventory_items "
    f"WHERE warehouse_id={q(MAIN_WH)} AND canonical_item_id = id AND retired_at IS NULL) < {n_can} "
    f"THEN RAISE EXCEPTION 'Phase 2: canonicals updated < expected {n_can}'; END IF; END $c$;")
add("SAVEPOINT after_canonicals;\n")

# ---------- Phase 3 — Alias retirement ----------
add("-- ------------------------------------------------------------------")
add("-- Phase 3 — Alias retirement (excludes FAT which stays independent)")
add("-- ------------------------------------------------------------------")
retired_aliases = []
for r in rows('Alias Retirement Plan'):
    sid = r['Source Item ID']
    cid = r['Canonical Item ID']
    if sid == FAT or cid == FAT: continue
    add(f"UPDATE public.inventory_items SET "
        f"canonical_item_id={q(cid)}, is_active=false, "
        f"retired_at=now(), retired_reason={q(STK)} "
        f"WHERE id={q(sid)} AND warehouse_id={q(MAIN_WH)};")
    retired_aliases.append(sid)
add(f"-- {len(retired_aliases)} alias rows retired.")
add("SAVEPOINT after_aliases;\n")

# ---------- Phase 4 — Product link rewires ----------
add("-- ------------------------------------------------------------------")
add(f"-- Phase 4 — Product Link Actions ({plan_rewire} rewire + {plan_keep} keep + {plan_independent} independent)")
add("-- ------------------------------------------------------------------")
n_rw = 0
for r in rows('Product Link Action Plan'):
    pid = r['Product ID']
    src = r['Current inventory_item_id (source)']
    can = r['Canonical Item ID']
    if src == FAT or can == FAT: continue
    if src == can: continue
    add(f"-- rewire product {pid}: {src} -> {can}")
    add(f"UPDATE public.inventory_items SET product_id=NULL "
        f"WHERE id={q(src)} AND warehouse_id={q(MAIN_WH)} AND product_id={q(pid)};")
    add(f"UPDATE public.inventory_items SET product_id={q(pid)} "
        f"WHERE id={q(can)} AND warehouse_id={q(MAIN_WH)} "
        f"AND (product_id IS NULL OR product_id={q(pid)});")
    n_rw += 1
add(f"-- {n_rw} product link rewires applied.")
add("SAVEPOINT after_product_links;\n")

# ---------- Phase 5 — Opening reversals (status flip). FAT skipped. ----------
add("-- ------------------------------------------------------------------")
add("-- Phase 5 — Duplicate opening reversals (status flip; NO new movement).")
add("--            FAT opening excluded — kept as-is in ledger (out of scope).")
add("-- ------------------------------------------------------------------")
n_rev = 0
for r in rows('Opening Reversal Actions'):
    mid  = r['Movement ID (to flip)']
    sid  = r.get('Source Item ID')
    kept = r['Kept Opening Movement ID']
    if sid == FAT: continue
    add(f"UPDATE public.inventory_movements SET "
        f"approval_status='reversed', reversed_at=now(), "
        f"reversed_by=current_setting('app.stk_actor')::uuid, "
        f"reversal_of_movement_id={q(kept)}, "
        f"reversal_reason={q('duplicate opening; kept=' + str(kept))}, "
        f"stocktake_ref={q(STK)} "
        f"WHERE id={q(mid)} AND warehouse_id={q(MAIN_WH)} AND approval_status='posted';")
    n_rev += 1
add(f"DO $r$ BEGIN IF (SELECT count(*) FROM public.inventory_movements "
    f"WHERE warehouse_id={q(MAIN_WH)} AND stocktake_ref={q(STK)} "
    f"AND approval_status='reversed') <> {n_rev} "
    f"THEN RAISE EXCEPTION 'Phase 5: expected {n_rev} reversed openings'; END IF; END $r$;")
add("SAVEPOINT after_openings;\n")

# ---------- Phase 6a — Reroute trigger (SCOPED to reclass reference_types only) ----------
add("""-- ------------------------------------------------------------------
-- Phase 6a — Reroute trigger. IMPORTANT SCOPE:
--   Rows whose reference_type is 'stocktake_reclass' or 'stocktake_reclass_neg'
--   are ALLOWED to land on the alias item_id as-written (that is the whole
--   point of reclassification legs). Every other insert — including future
--   stocktake_adjustment rows and normal operations — still routes to the
--   canonical. This narrower rule replaces v5.1's blanket
--   "stocktake_ref IS NOT NULL" bypass which would have leaked to future
--   stocktakes.
-- ------------------------------------------------------------------
CREATE OR REPLACE FUNCTION public.reroute_to_canonical()
RETURNS trigger LANGUAGE plpgsql SET search_path = public AS $fn$
DECLARE can uuid;
BEGIN
  IF NEW.reference_type IN ('stocktake_reclass','stocktake_reclass_neg') THEN
    RETURN NEW;  -- reclassification legs are the exception
  END IF;
  SELECT canonical_item_id INTO can FROM public.inventory_items WHERE id = NEW.item_id;
  IF can IS NOT NULL AND can <> NEW.item_id THEN NEW.item_id := can; END IF;
  RETURN NEW;
END $fn$;

CREATE TRIGGER trg_inv_mov_reroute_canonical BEFORE INSERT ON public.inventory_movements
FOR EACH ROW EXECUTE FUNCTION public.reroute_to_canonical();

-- Reservation reroute is intentionally SCOPED to agouza_stock_reservations,
-- which is the only reservations table with inventory_item_id tied to the
-- main warehouse in the current schema. Other modules will be handled
-- individually if/when they receive alias-based inventory FKs.
CREATE OR REPLACE FUNCTION public.reroute_res_to_canonical()
RETURNS trigger LANGUAGE plpgsql SET search_path = public AS $fn$
DECLARE can uuid;
BEGIN
  SELECT canonical_item_id INTO can FROM public.inventory_items WHERE id = NEW.inventory_item_id;
  IF can IS NOT NULL AND can <> NEW.inventory_item_id THEN NEW.inventory_item_id := can; END IF;
  RETURN NEW;
END $fn$;

CREATE TRIGGER trg_agouza_res_reroute BEFORE INSERT ON public.agouza_stock_reservations
FOR EACH ROW EXECUTE FUNCTION public.reroute_res_to_canonical();
SAVEPOINT after_triggers;
""")

# ---------- Phase 6b — 126 movements (fixed neg-reclass IN on alias) ----------
add("-- ------------------------------------------------------------------")
add("-- Phase 6b — 126 posted movements (one INSERT per Movement Action Plan row)")
add("--   RECLASSIFY:          OUT lands on ALIAS,     IN lands on CANONICAL")
add("--   RECLASSIFY_NEGATIVE: OUT lands on CANONICAL, IN lands on ALIAS   ← fixed vs v5.1")
add("--   STOCKTAKE_ADJUSTMENT: lands on CANONICAL")
add("-- ------------------------------------------------------------------")
n_reclass = n_adjust = 0
for r in rows('Movement Action Plan'):
    op        = r['Operation Type']
    sid       = r['Source Item ID']
    cid       = r['Canonical Item ID']
    qty       = r['Quantity']
    direction = str(r['Direction'] or '')
    idem      = r['Idempotency Key']
    ref_type  = r['Reference Type']
    notes     = r['notes']
    is_out    = direction.startswith('OUT')
    db_mt     = 'out' if is_out else 'in'
    if op == 'RECLASSIFY':
        # alias→canonical: OUT alias / IN canonical
        item_id = sid if is_out else cid
        n_reclass += 1
    elif op == 'RECLASSIFY_NEGATIVE':
        # canonical→alias: OUT canonical / IN alias   ← FIX
        item_id = cid if is_out else sid
        n_reclass += 1
    else:  # STOCKTAKE_ADJUSTMENT
        item_id = cid
        n_adjust += 1
    add(f"-- {op} {direction} qty={qty} item={item_id}")
    add(f"INSERT INTO public.inventory_movements "
        f"(item_id, warehouse_id, movement_type, quantity, unit_cost, notes, "
        f" reference_type, reference_id, stocktake_ref, idempotency_key, "
        f" approval_status, performed_by, approved_by, approved_at) "
        f"VALUES ({q(item_id)}, {q(MAIN_WH)}, {q(db_mt)}, {q(qty)}, 0, {q(notes)}, "
        f"       {q(ref_type)}, {q(STK)}, {q(STK)}, {q(idem)}, "
        f"       'posted', current_setting('app.stk_actor')::uuid, "
        f"       current_setting('app.stk_actor')::uuid, now());")

add(f"-- reclassification legs inserted: {n_reclass} (expected 100)")
add(f"-- stocktake adjustments inserted: {n_adjust} (expected 26)")
add(f"DO $m$ DECLARE c int; BEGIN "
    f"SELECT count(*) INTO c FROM public.inventory_movements "
    f"WHERE warehouse_id={q(MAIN_WH)} AND stocktake_ref={q(STK)} AND idempotency_key IS NOT NULL; "
    f"IF c <> {n_reclass + n_adjust} THEN "
    f"  RAISE EXCEPTION 'Phase 6b: inserted % movements, expected {n_reclass + n_adjust}', c; "
    f"END IF; END $m$;")
add("SAVEPOINT after_movements;\n")

# ---------- Phase 7 — Stocktake session + 27 lines with correct system_qty/actual_qty ----------
# Build lookup: canonical_id -> (system_qty, actual_qty) from Adjustments
adj_map = {}
for r in rows('Adjustments'):
    cid = r.get('Canonical ID')
    if not cid: continue
    adj_map[cid] = (
        r.get('Canonical Book After Reversal+Reclass') or 0,   # system_qty
        r.get('Actual (Physical Count)') or 0,                  # actual_qty
    )

add(f"""-- ------------------------------------------------------------------
-- Phase 7 — Stocktake session header + 27 stocktaking_lines
--            system_qty = Canonical Book After Reversal+Reclass (pre-adjust)
--            actual_qty = Physical Count
-- ------------------------------------------------------------------
WITH s AS (
  INSERT INTO public.stocktaking_sessions
    (session_no, warehouse_id, count_date, stocktaker_name, status, notes, created_by, approved_by, approved_at, reference_id)
  VALUES ({q(STK)}, {q(MAIN_WH)}, DATE '2026-07-27',
          'Physical Count Team', 'approved',
          'MAIN-STOCKTAKE-20260727: 26 kg items totalling 624.5 kg + 6ك دبوس بالعظم adjusted to 0. FAT pack out of scope.',
          current_setting('app.stk_actor')::uuid,
          current_setting('app.stk_actor')::uuid, now(), {q(STK)})
  RETURNING id
)
INSERT INTO public.stocktaking_lines (session_id, item_id, system_qty, actual_qty, unit_cost, reason, notes)
SELECT s.id, v.item_id, v.system_qty, v.actual_qty, 0, v.reason, v.notes
FROM s, (VALUES""")

line_vals = []
for cid in canonicals_all:
    sys_qty, act_qty = adj_map.get(cid, (0, 0))
    # get final name for note
    fname = ''
    for r in rows('Item Master Action Plan'):
        if r['Canonical Item ID'] == cid:
            fname = r['Final Name'] or ''
            break
    line_vals.append(f"({q(cid)}::uuid, {q(sys_qty)}, {q(act_qty)}, {q('stocktake_' + STK)}, {q(fname)})")
add(",\n".join(line_vals))
add(") AS v(item_id, system_qty, actual_qty, reason, notes);")

# assertion: 27 lines exist
add(f"DO $sl$ DECLARE c int; BEGIN "
    f"SELECT count(*) INTO c FROM public.stocktaking_lines l "
    f"JOIN public.stocktaking_sessions s ON s.id=l.session_id "
    f"WHERE s.session_no={q(STK)}; "
    f"IF c <> {len(canonicals_all)} THEN "
    f"  RAISE EXCEPTION 'Phase 7: stocktaking_lines count %, expected {len(canonicals_all)}', c; "
    f"END IF; END $sl$;")
add("SAVEPOINT after_session;\n")

# ---------- Phase 8 — Views (security_invoker) ----------
add(f"""-- ------------------------------------------------------------------
-- Phase 8 — Views. Canonical grouping = one row per (canonical, warehouse).
-- ------------------------------------------------------------------
CREATE VIEW public.v_inventory_balances
WITH (security_invoker = true) AS
SELECT
  COALESCE(i.canonical_item_id, m.item_id) AS canonical_item_id,
  m.warehouse_id,
  SUM(CASE
        WHEN m.movement_type IN ('out','sales_dispatch','stock_out',
                                 'production_consumption','packaging_consumption','waste_loss')
          THEN -m.quantity ELSE m.quantity END) AS balance
FROM public.inventory_movements m
LEFT JOIN public.inventory_items i ON i.id = m.item_id
WHERE m.approval_status = 'posted'
GROUP BY 1, 2;
GRANT SELECT ON public.v_inventory_balances TO authenticated, service_role;

CREATE VIEW public.v_inventory_balances_raw
WITH (security_invoker = true) AS
SELECT
  m.item_id AS raw_item_id,
  COALESCE(i.canonical_item_id, m.item_id) AS canonical_item_id,
  m.warehouse_id,
  SUM(CASE
        WHEN m.movement_type IN ('out','sales_dispatch','stock_out',
                                 'production_consumption','packaging_consumption','waste_loss')
          THEN -m.quantity ELSE m.quantity END) AS balance
FROM public.inventory_movements m
LEFT JOIN public.inventory_items i ON i.id = m.item_id
WHERE m.approval_status = 'posted'
GROUP BY 1, 2, 3;
GRANT SELECT ON public.v_inventory_balances_raw TO authenticated, service_role;
SAVEPOINT after_view;

-- ------------------------------------------------------------------
-- Phase 9 — Sync inventory_items.stock from canonical view; zero retired aliases
-- ------------------------------------------------------------------
UPDATE public.inventory_items i
SET stock = COALESCE(v.balance, 0)
FROM public.v_inventory_balances v
WHERE i.id = v.canonical_item_id
  AND i.warehouse_id = '{MAIN_WH}'
  AND v.warehouse_id = '{MAIN_WH}'
  AND i.canonical_item_id = i.id
  AND i.id <> '{FAT}';

UPDATE public.inventory_items SET stock = 0
WHERE warehouse_id = '{MAIN_WH}' AND retired_at IS NOT NULL;
""")

# ---------- Phase 10 — Assertions ----------
retired_list_sql = ",".join(f"'{sid}'" for sid in retired_aliases)
kg_ids_sql = ",".join(f"'{c}'" for c in canonicals_kg)
add(f"""-- ------------------------------------------------------------------
-- Phase 10 — Assertions (any failure aborts the whole transaction)
-- ------------------------------------------------------------------
DO $chk$
DECLARE
  kg_total       numeric;
  pin6k_bal      numeric;
  dup_canonical  int;
  other_wh_count bigint;
  snap_other     bigint;
  fat_touched    int;
  bad_alias      int;
BEGIN
  -- 10.1 kg subtotal (26 kg canonicals only)
  SELECT COALESCE(SUM(v.balance),0) INTO kg_total
  FROM public.v_inventory_balances v
  WHERE v.warehouse_id = '{MAIN_WH}'
    AND v.canonical_item_id IN ({kg_ids_sql});
  IF ROUND(kg_total::numeric,2) <> 624.50 THEN
    RAISE EXCEPTION 'Assertion 10.1: kg total = %, expected 624.50', kg_total;
  END IF;

  -- 10.2 6ك دبوس بالعظم canonical = 0
  SELECT COALESCE(SUM(balance),0) INTO pin6k_bal
  FROM public.v_inventory_balances
  WHERE canonical_item_id = '{PIN_6K}' AND warehouse_id = '{MAIN_WH}';
  IF ROUND(pin6k_bal::numeric,4) <> 0 THEN
    RAISE EXCEPTION 'Assertion 10.2: 6ك دبوس بالعظم balance = %, expected 0', pin6k_bal;
  END IF;

  -- 10.3 View returns exactly one row per canonical in main WH
  SELECT count(*) INTO dup_canonical FROM (
    SELECT canonical_item_id FROM public.v_inventory_balances
    WHERE warehouse_id = '{MAIN_WH}'
    GROUP BY canonical_item_id HAVING count(*) > 1
  ) d;
  IF dup_canonical <> 0 THEN
    RAISE EXCEPTION 'Assertion 10.3: v_inventory_balances has % duplicated canonicals', dup_canonical;
  END IF;

  -- 10.4 Other-warehouse item count unchanged
  SELECT count(*) INTO other_wh_count FROM public.inventory_items WHERE warehouse_id <> '{MAIN_WH}';
  SELECT row_count INTO snap_other FROM public.stocktake_20260727_checksums WHERE label='items_before_other_wh';
  IF other_wh_count <> snap_other THEN
    RAISE EXCEPTION 'Assertion 10.4: other-warehouse item count changed (% vs %)', other_wh_count, snap_other;
  END IF;

  -- 10.5 FAT pack untouched by this migration
  SELECT count(*) INTO fat_touched
  FROM public.inventory_movements
  WHERE item_id = '{FAT}' AND stocktake_ref = '{STK}';
  IF fat_touched <> 0 THEN
    RAISE EXCEPTION 'Assertion 10.5: FAT pack was touched by % movements in this stocktake', fat_touched;
  END IF;

  -- 10.6 Every retired alias's RAW ledger balance = 0
  SELECT count(*) INTO bad_alias
  FROM public.v_inventory_balances_raw r
  WHERE r.warehouse_id = '{MAIN_WH}'
    AND r.raw_item_id IN ({retired_list_sql if retired_list_sql else "NULL"})
    AND COALESCE(ROUND(r.balance::numeric,4),0) <> 0;
  IF bad_alias <> 0 THEN
    RAISE EXCEPTION 'Assertion 10.6: % retired alias(es) have non-zero raw ledger balance', bad_alias;
  END IF;
END $chk$;

-- After-state checksums
INSERT INTO public.stocktake_20260727_checksums(label,row_count,qty_sum,name_unit_sku_hash)
SELECT 'items_after_main', count(*)::bigint, sum(stock),
       md5(string_agg(coalesce(name,'')||'|'||coalesce(unit,'')||'|'||coalesce(sku,''), '||' ORDER BY id))
FROM public.inventory_items WHERE warehouse_id='{MAIN_WH}';
INSERT INTO public.stocktake_20260727_checksums(label,row_count,qty_sum)
SELECT 'movements_after_main', count(*)::bigint, sum(quantity)
FROM public.inventory_movements WHERE warehouse_id='{MAIN_WH}';
INSERT INTO public.stocktake_20260727_checksums(label,row_count,qty_sum,name_unit_sku_hash)
SELECT 'items_after_other_wh', count(*)::bigint, sum(stock),
       md5(string_agg(coalesce(name,'')||'|'||coalesce(unit,'')||'|'||coalesce(sku,''), '||' ORDER BY id))
FROM public.inventory_items WHERE warehouse_id<>'{MAIN_WH}';

COMMIT;
-- ============================================================================
-- End of MAIN-STOCKTAKE-20260727 forward migration v5.2
-- ============================================================================
""")

with open(FWD,'w') as f: f.write('\n'.join(out))
print('WROTE', FWD)

# ============================================================================
# ROLLBACK SQL v5.2 — explicit columns + full checksum verification
# ============================================================================
r = []
r.append(f"""-- ============================================================================
-- MAIN-STOCKTAKE-20260727 — Rollback v5.2
-- Requires:  SELECT set_config('app.stk_actor','<uuid>', false);   -- BEFORE run
-- No psql meta commands.
-- Verifies: row counts + sum(stock) + name|unit|sku hash vs pre-forward snapshot.
-- ============================================================================
BEGIN;
SET LOCAL statement_timeout = '15min';
SET LOCAL lock_timeout      = '30s';

DO $g$
DECLARE v_actor uuid; ok boolean := false; v_setting text;
BEGIN
  BEGIN v_setting := current_setting('app.stk_actor', true); EXCEPTION WHEN OTHERS THEN v_setting := NULL; END;
  v_actor := COALESCE(auth.uid(), NULLIF(v_setting,'')::uuid);
  IF v_actor IS NULL THEN RAISE EXCEPTION 'Rollback: no actor. SELECT set_config(''app.stk_actor'',''<uuid>'',false) before running.'; END IF;
  PERFORM set_config('app.stk_actor', v_actor::text, true);
  SELECT bool_or(role::text IN ('general_manager','executive_manager')) INTO ok
  FROM public.user_roles WHERE user_id = v_actor;
  IF NOT COALESCE(ok,false) THEN RAISE EXCEPTION 'Rollback: actor lacks GM/Executive role'; END IF;
END $g$;

LOCK TABLE public.inventory_items     IN SHARE ROW EXCLUSIVE MODE;
LOCK TABLE public.inventory_movements IN SHARE ROW EXCLUSIVE MODE;

-- 1) Remove movements INSERTED by the migration
DELETE FROM public.inventory_movements
WHERE warehouse_id = '{MAIN_WH}'
  AND stocktake_ref = '{STK}'
  AND idempotency_key IS NOT NULL;

-- 2) Restore inventory_movements rows (status flips + any updated columns)
INSERT INTO public.inventory_movements AS m ({im_cols})
SELECT {im_cols} FROM public.inventory_movements_snapshot_20260727 s
ON CONFLICT (id) DO UPDATE SET
""")
set_cols = [c for c in IM_COLS if c != 'id']
r.append(',\n'.join(f'  {c} = EXCLUDED.{c}' for c in set_cols) + ';')

r.append(f"""
UPDATE public.inventory_movements
SET reversal_of_movement_id = NULL, reversed_at = NULL, reversed_by = NULL,
    reversal_reason = NULL, stocktake_ref = NULL
WHERE stocktake_ref = '{STK}';

-- 3) Restore inventory_items rows
INSERT INTO public.inventory_items AS i ({ii_cols})
SELECT {ii_cols} FROM public.inventory_items_snapshot_20260727 s
ON CONFLICT (id) DO UPDATE SET
""")
set_cols_i = [c for c in II_COLS if c != 'id']
r.append(',\n'.join(f'  {c} = EXCLUDED.{c}' for c in set_cols_i) + ';')

r.append(f"""
UPDATE public.inventory_items
SET canonical_item_id = NULL, retired_at = NULL, retired_reason = NULL
WHERE warehouse_id = '{MAIN_WH}';

-- 4) Remove stocktake session lines + header
DELETE FROM public.stocktaking_lines
WHERE session_id IN (SELECT id FROM public.stocktaking_sessions
                     WHERE session_no = '{STK}' OR reference_id = '{STK}');
DELETE FROM public.stocktaking_sessions
WHERE session_no = '{STK}' OR reference_id = '{STK}';

-- 5) Drop schema objects created by forward
DROP TRIGGER  IF EXISTS trg_inv_mov_reroute_canonical ON public.inventory_movements;
DROP TRIGGER  IF EXISTS trg_agouza_res_reroute        ON public.agouza_stock_reservations;
DROP FUNCTION IF EXISTS public.reroute_to_canonical();
DROP FUNCTION IF EXISTS public.reroute_res_to_canonical();
DROP VIEW     IF EXISTS public.v_inventory_balances_raw;
DROP VIEW     IF EXISTS public.v_inventory_balances;
DROP INDEX    IF EXISTS public.ux_inv_mov_idempotency_key;

ALTER TABLE public.inventory_movements
  DROP COLUMN IF EXISTS idempotency_key,
  DROP COLUMN IF EXISTS reversal_of_movement_id,
  DROP COLUMN IF EXISTS reversed_at,
  DROP COLUMN IF EXISTS reversed_by,
  DROP COLUMN IF EXISTS reversal_reason,
  DROP COLUMN IF EXISTS stocktake_ref;
ALTER TABLE public.inventory_items
  DROP COLUMN IF EXISTS canonical_item_id,
  DROP COLUMN IF EXISTS retired_at,
  DROP COLUMN IF EXISTS retired_reason;

-- 6) VERIFY: count + sum(stock) + name|unit|sku hash match pre-forward snapshot
DO $v$
DECLARE
  want_rc bigint; got_rc bigint;
  want_qty numeric; got_qty numeric;
  want_h text; got_h text;
BEGIN
  -- items main WH
  SELECT row_count, qty_sum, name_unit_sku_hash INTO want_rc, want_qty, want_h
  FROM public.stocktake_20260727_checksums WHERE label='items_before_main';
  SELECT count(*)::bigint, sum(stock),
         md5(string_agg(coalesce(name,'')||'|'||coalesce(unit,'')||'|'||coalesce(sku,''), '||' ORDER BY id))
  INTO got_rc, got_qty, got_h
  FROM public.inventory_items WHERE warehouse_id='{MAIN_WH}';
  IF want_rc <> got_rc THEN RAISE EXCEPTION 'Rollback: items count mismatch want=% got=%', want_rc, got_rc; END IF;
  IF ROUND(COALESCE(want_qty,0),4) <> ROUND(COALESCE(got_qty,0),4) THEN
    RAISE EXCEPTION 'Rollback: items sum(stock) mismatch want=% got=%', want_qty, got_qty;
  END IF;
  IF want_h IS DISTINCT FROM got_h THEN
    RAISE EXCEPTION 'Rollback: items name|unit|sku hash mismatch (schema drift or data changed)';
  END IF;

  -- movements main WH
  SELECT row_count, qty_sum INTO want_rc, want_qty
  FROM public.stocktake_20260727_checksums WHERE label='movements_before_main';
  SELECT count(*)::bigint, sum(quantity) INTO got_rc, got_qty
  FROM public.inventory_movements WHERE warehouse_id='{MAIN_WH}';
  IF want_rc <> got_rc THEN RAISE EXCEPTION 'Rollback: movements count mismatch want=% got=%', want_rc, got_rc; END IF;
  IF ROUND(COALESCE(want_qty,0),4) <> ROUND(COALESCE(got_qty,0),4) THEN
    RAISE EXCEPTION 'Rollback: movements sum(quantity) mismatch want=% got=%', want_qty, got_qty;
  END IF;

  -- other warehouses (must not have changed)
  SELECT row_count, qty_sum, name_unit_sku_hash INTO want_rc, want_qty, want_h
  FROM public.stocktake_20260727_checksums WHERE label='items_before_other_wh';
  SELECT count(*)::bigint, sum(stock),
         md5(string_agg(coalesce(name,'')||'|'||coalesce(unit,'')||'|'||coalesce(sku,''), '||' ORDER BY id))
  INTO got_rc, got_qty, got_h
  FROM public.inventory_items WHERE warehouse_id<>'{MAIN_WH}';
  IF want_rc <> got_rc OR want_h IS DISTINCT FROM got_h THEN
    RAISE EXCEPTION 'Rollback: OTHER warehouses drift detected (should have been untouched)';
  END IF;
END $v$;

COMMIT;

-- Snapshot & checksum tables remain until you drop manually:
--   DROP TABLE public.inventory_items_snapshot_20260727;
--   DROP TABLE public.inventory_movements_snapshot_20260727;
--   DROP TABLE public.stocktake_20260727_checksums;
""")

with open(RBK,'w') as f: f.write('\n'.join(r))
print('WROTE', RBK)
