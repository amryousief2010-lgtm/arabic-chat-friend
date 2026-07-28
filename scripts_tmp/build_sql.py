"""Generate forward_migration + rollback SQL from DryRun v5 workbook."""
import openpyxl, json, textwrap
from datetime import datetime

V5 = '/mnt/documents/DryRun_MAIN-STOCKTAKE-20260727_v5.xlsx'
FWD = '/mnt/documents/forward_migration_MAIN-STOCKTAKE-20260727.sql'
RBK = '/mnt/documents/rollback_MAIN-STOCKTAKE-20260727.sql'
STK = 'MAIN-STOCKTAKE-20260727'
MAIN_WH = '5ec781b5-685b-4806-b59a-83a79ea5662c'   # المخزن الرئيسي - المقر
FAT = 'a55763c7-6e12-4ed0-8d62-37ace6bf86a3'

wb = openpyxl.load_workbook(V5)

def rows(sheet):
    ws = wb[sheet]
    hdr = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None: continue
        yield dict(zip(hdr, row))

def q(v):
    if v is None: return 'NULL'
    if isinstance(v, bool): return 'true' if v else 'false'
    if isinstance(v, (int, float)): return str(v)
    return "'" + str(v).replace("'", "''") + "'"

# ---------- FORWARD MIGRATION ----------
out = []
add = out.append

add(f"""-- ============================================================================
-- Forward Migration: {STK}
-- Scope: المخزن الرئيسي - المقر ONLY (warehouse_id = {MAIN_WH})
-- Generated: {datetime.utcnow().isoformat()}Z from DryRun v5
-- Method: status-flip reversals + alias→canonical rewire + stocktake session
-- Safety: single BEGIN/COMMIT + SAVEPOINTs + assertions; no DELETE/TRUNCATE/DROP
-- Execution: DO NOT RUN without written approval. Review this file end-to-end.
-- ============================================================================
BEGIN;
SET LOCAL statement_timeout = '10min';
SET LOCAL lock_timeout      = '30s';

-- Guard: must be executed by GM/Executive
DO $mig$
DECLARE r text; ok boolean := false;
BEGIN
  FOR r IN SELECT role::text FROM public.user_roles WHERE user_id = auth.uid()
  LOOP IF r IN ('general_manager','executive_manager') THEN ok := true; END IF; END LOOP;
  IF NOT ok THEN RAISE EXCEPTION 'Only GM/Executive can execute {STK}'; END IF;
END $mig$;

-- ------------------------------------------------------------------
-- Phase 0 — Snapshots (main warehouse rows only, full row copy)
-- ------------------------------------------------------------------
CREATE TABLE public.inventory_items_snapshot_20260727 AS
SELECT * FROM public.inventory_items WHERE warehouse_id = '{MAIN_WH}';
ALTER TABLE public.inventory_items_snapshot_20260727
  ADD CONSTRAINT inventory_items_snapshot_20260727_pk PRIMARY KEY (id);

CREATE TABLE public.inventory_movements_snapshot_20260727 AS
SELECT * FROM public.inventory_movements WHERE warehouse_id = '{MAIN_WH}';
ALTER TABLE public.inventory_movements_snapshot_20260727
  ADD CONSTRAINT inventory_movements_snapshot_20260727_pk PRIMARY KEY (id);

-- Checksum baselines
CREATE TABLE public.stocktake_20260727_checksums (
  label text PRIMARY KEY,
  row_count bigint NOT NULL,
  qty_sum numeric,
  captured_at timestamptz NOT NULL DEFAULT now()
);
INSERT INTO public.stocktake_20260727_checksums(label,row_count,qty_sum)
SELECT 'items_before', count(*)::bigint, NULL FROM public.inventory_items_snapshot_20260727;
INSERT INTO public.stocktake_20260727_checksums(label,row_count,qty_sum)
SELECT 'movements_before', count(*)::bigint,
       (SELECT sum(CASE WHEN movement_type IN ('out','sales_dispatch','stock_out','production_consumption','packaging_consumption','waste_loss') THEN -quantity ELSE quantity END)
        FROM public.inventory_movements_snapshot_20260727 WHERE approval_status='posted')
FROM public.inventory_movements_snapshot_20260727;

SAVEPOINT after_snapshot;

-- ------------------------------------------------------------------
-- Phase 1 — Schema additions (idempotent guards, canonical/alias infra)
-- ------------------------------------------------------------------
ALTER TABLE public.inventory_items
  ADD COLUMN IF NOT EXISTS canonical_item_id uuid REFERENCES public.inventory_items(id),
  ADD COLUMN IF NOT EXISTS retired_at        timestamptz,
  ADD COLUMN IF NOT EXISTS retired_reason    text;

ALTER TABLE public.inventory_movements
  ADD COLUMN IF NOT EXISTS idempotency_key        text,
  ADD COLUMN IF NOT EXISTS reversal_of_movement_id uuid REFERENCES public.inventory_movements(id),
  ADD COLUMN IF NOT EXISTS reversed_at            timestamptz,
  ADD COLUMN IF NOT EXISTS reversed_by            uuid REFERENCES auth.users(id),
  ADD COLUMN IF NOT EXISTS reversal_reason        text,
  ADD COLUMN IF NOT EXISTS stocktake_ref          text;

CREATE UNIQUE INDEX IF NOT EXISTS ux_inv_mov_idempotency_key
  ON public.inventory_movements(idempotency_key)
  WHERE idempotency_key IS NOT NULL;

SAVEPOINT after_schema;
""")

# ------------------------------------------------------------------
# Phase 2 — Item Master (canonicals + aliases)
# ------------------------------------------------------------------
add("-- ------------------------------------------------------------------")
add("-- Phase 2 — Item Master: update 27 canonicals; retire 80 aliases")
add("--                       Fat pack (a557…) stays independent (canonical=self)")
add("-- ------------------------------------------------------------------")

# canonicals from Item Master Action Plan (each row is a canonical)
n_can = 0
for r in rows('Item Master Action Plan'):
    cid   = r['Canonical Item ID']
    fname = r['Final Name']
    funit = r['Final Unit']
    fbc   = r['Final Barcode (text)']
    add(f"UPDATE public.inventory_items SET "
        f"name={q(fname)}, unit={q(funit)}, "
        f"sku=COALESCE(sku,{q(fbc)}), "
        f"canonical_item_id=id, "
        f"is_active=true, retired_at=NULL "
        f"WHERE id={q(cid)} AND warehouse_id={q(MAIN_WH)};")
    n_can += 1
add(f"-- {n_can} canonical rows updated (expected 27).")
add("SAVEPOINT after_canonicals;\n")

# aliases from Alias Retirement Plan
n_al = 0
for r in rows('Alias Retirement Plan'):
    sid = r['Source Item ID']
    cid = r['Canonical Item ID']
    if sid == FAT:  # safety
        continue
    add(f"UPDATE public.inventory_items SET "
        f"canonical_item_id={q(cid)}, "
        f"is_active=false, "
        f"retired_at=now(), retired_reason={q(STK)} "
        f"WHERE id={q(sid)} AND warehouse_id={q(MAIN_WH)};")
    n_al += 1
add(f"-- {n_al} alias rows retired (expected 80).")
add("SAVEPOINT after_aliases;\n")

# ------------------------------------------------------------------
# Phase 3 — Product link rewires (products.inventory_item_id lives on inventory_items.product_id)
# ------------------------------------------------------------------
add("-- ------------------------------------------------------------------")
add("-- Phase 3 — Product link rewires: move product_id from alias→canonical")
add("-- Two-step to respect UNIQUE(warehouse_id, product_id) partial index.")
add("-- ------------------------------------------------------------------")
rewires = []
for r in rows('Product Link Action Plan'):
    pid = r['Product ID']
    src = r['Current inventory_item_id (source)']
    can = r['Canonical Item ID']
    if src == FAT or can == FAT or r['Display-Only Alias?'] is True:
        # fat pack: keep own link, no rewire
        continue
    if src == can:
        # already canonical → nothing to do
        continue
    rewires.append((pid, src, can))

for pid, src, can in rewires:
    add(f"-- product {pid}: {src} → {can}")
    add(f"UPDATE public.inventory_items SET product_id=NULL "
        f"WHERE id={q(src)} AND warehouse_id={q(MAIN_WH)} AND product_id={q(pid)};")
    add(f"UPDATE public.inventory_items SET product_id={q(pid)} "
        f"WHERE id={q(can)} AND warehouse_id={q(MAIN_WH)} "
        f"AND (product_id IS NULL OR product_id={q(pid)});")
add(f"-- {len(rewires)} product link rewires applied.")
add("SAVEPOINT after_product_links;\n")

# ------------------------------------------------------------------
# Phase 4 — Opening reversals (status flip on 104 original rows)
# ------------------------------------------------------------------
add("-- ------------------------------------------------------------------")
add("-- Phase 4 — 104 duplicate opening reversals (status flip; NO new movement)")
add("-- ------------------------------------------------------------------")
n_rev = 0
for r in rows('Opening Reversal Actions'):
    mid = r['Movement ID (to flip)']
    kept = r['Kept Opening Movement ID']
    add(f"UPDATE public.inventory_movements SET "
        f"approval_status='reversed', reversed_at=now(), reversed_by=auth.uid(), "
        f"reversal_of_movement_id={q(kept)}, "
        f"reversal_reason={q('duplicate opening; kept=' + str(kept))}, "
        f"stocktake_ref={q(STK)} "
        f"WHERE id={q(mid)} AND warehouse_id={q(MAIN_WH)} AND approval_status='posted';")
    n_rev += 1
add(f"-- {n_rev} openings flipped (expected 104).")
add("SAVEPOINT after_openings;\n")

# ------------------------------------------------------------------
# Phase 5 — Reclassification (alias→canonical, INTERNAL to main warehouse)
# ------------------------------------------------------------------
add("-- ------------------------------------------------------------------")
add("-- Phase 5 — Reclassification: net qty transferred from alias to canonical")
add("-- movement_type='adjustment' (existing CHECK-allowed value); idempotency_key")
add("-- enforces one-time application.")
add("-- ------------------------------------------------------------------")
n_re = 0
for r in rows('Movement Action Plan'):
    op = r['Operation Type']
    if op not in ('RECLASSIFY','RECLASSIFY_NEGATIVE'):
        continue
    if op == 'STOCKTAKE_ADJUSTMENT':
        continue
    sid = r['Source Item ID']
    cid = r['Canonical Item ID']
    qty = r['Quantity']
    unit = r['Unit']
    direction = r['Direction']
    idem = r['Idempotency Key']
    notes = r['notes']
    # OUT from alias, IN to canonical — same qty. Net zero on main warehouse total.
    # Direction wording differs between RECLASSIFY (alias→canonical) and RECLASSIFY_NEGATIVE (repair negatives).
    if 'OUT' in str(direction):
        mt_out, mt_in = 'out', 'in'
    else:
        mt_out, mt_in = 'in', 'out'  # negative case: canonical→alias to zero it out
    add(f"-- {op}: {sid} → {cid} qty={qty} {unit}  ({notes})")
    add(f"INSERT INTO public.inventory_movements "
        f"(item_id, warehouse_id, movement_type, quantity, unit_cost, notes, "
        f" reference_type, reference_id, stocktake_ref, idempotency_key, approval_status, performed_by) "
        f"VALUES ({q(sid)}, {q(MAIN_WH)}, {q(mt_out)}, {q(qty)}, 0, {q('stocktake_reclass_out')}, "
        f"       'stocktake_reclass', {q(STK)}, {q(STK)}, {q(idem+':O')}, 'posted', auth.uid()) "
        f"ON CONFLICT (idempotency_key) WHERE idempotency_key IS NOT NULL DO NOTHING;")
    add(f"INSERT INTO public.inventory_movements "
        f"(item_id, warehouse_id, movement_type, quantity, unit_cost, notes, "
        f" reference_type, reference_id, stocktake_ref, idempotency_key, approval_status, performed_by) "
        f"VALUES ({q(cid)}, {q(MAIN_WH)}, {q(mt_in)}, {q(qty)}, 0, {q('stocktake_reclass_in')}, "
        f"       'stocktake_reclass', {q(STK)}, {q(STK)}, {q(idem+':I')}, 'posted', auth.uid()) "
        f"ON CONFLICT (idempotency_key) WHERE idempotency_key IS NOT NULL DO NOTHING;")
    n_re += 1
add(f"-- {n_re} reclassification pairs applied.")
add("SAVEPOINT after_reclass;\n")

# ------------------------------------------------------------------
# Phase 6 — Stocktake session + 26 adjustments
# ------------------------------------------------------------------
add("-- ------------------------------------------------------------------")
add("-- Phase 6 — Stocktake session and 26 adjustment lines")
add("-- ------------------------------------------------------------------")
add(f"INSERT INTO public.stocktaking_sessions(reference, warehouse_id, status, notes, created_by, created_at) "
    f"VALUES ({q(STK)}, {q(MAIN_WH)}, 'approved', 'Physical count 2026-07-27; 26 kg items totalling 624.5kg', auth.uid(), now()) "
    f"ON CONFLICT (reference) DO NOTHING;")
n_adj = 0
for r in rows('Movement Action Plan'):
    if r['Operation Type'] != 'STOCKTAKE_ADJUSTMENT': continue
    cid = r['Canonical Item ID']
    qty = r['Quantity']
    unit = r['Unit']
    idem = r['Idempotency Key']
    notes = r['notes']
    direction = str(r['Direction'])
    mt = 'out' if 'OUT' in direction else 'in'
    add(f"-- STOCKTAKE_ADJUSTMENT: {cid} {direction} {qty} {unit}  ({notes})")
    add(f"INSERT INTO public.inventory_movements "
        f"(item_id, warehouse_id, movement_type, quantity, unit_cost, notes, "
        f" reference_type, reference_id, stocktake_ref, idempotency_key, approval_status, performed_by) "
        f"VALUES ({q(cid)}, {q(MAIN_WH)}, {q(mt)}, {q(qty)}, 0, {q(notes)}, "
        f"       'stocktaking_session', {q(STK)}, {q(STK)}, {q(idem)}, 'posted', auth.uid()) "
        f"ON CONFLICT (idempotency_key) WHERE idempotency_key IS NOT NULL DO NOTHING;")
    n_adj += 1
add(f"-- {n_adj} stocktake adjustments applied (expected 26).")
add("SAVEPOINT after_adjustments;\n")

# ------------------------------------------------------------------
# Phase 7 — View + triggers
# ------------------------------------------------------------------
add("""-- ------------------------------------------------------------------
-- Phase 7 — Balance view + alias→canonical redirect triggers
-- ------------------------------------------------------------------
CREATE OR REPLACE VIEW public.v_inventory_balances AS
SELECT
  COALESCE(m.item_id, m.item_id)                       AS raw_item_id,
  COALESCE(i.canonical_item_id, m.item_id)             AS canonical_item_id,
  m.warehouse_id,
  SUM(CASE
        WHEN m.movement_type IN ('out','sales_dispatch','stock_out','production_consumption','packaging_consumption','waste_loss')
          THEN -m.quantity
        ELSE m.quantity
      END)                                             AS balance
FROM public.inventory_movements m
LEFT JOIN public.inventory_items i ON i.id = m.item_id
WHERE m.approval_status = 'posted'
GROUP BY 1,2,3;
GRANT SELECT ON public.v_inventory_balances TO authenticated;

CREATE OR REPLACE FUNCTION public.reroute_to_canonical()
RETURNS trigger LANGUAGE plpgsql SET search_path = public AS $fn$
DECLARE can uuid;
BEGIN
  SELECT canonical_item_id INTO can FROM public.inventory_items WHERE id = NEW.item_id;
  IF can IS NOT NULL AND can <> NEW.item_id THEN NEW.item_id := can; END IF;
  RETURN NEW;
END $fn$;

DROP TRIGGER IF EXISTS trg_inv_mov_reroute_canonical ON public.inventory_movements;
CREATE TRIGGER trg_inv_mov_reroute_canonical BEFORE INSERT ON public.inventory_movements
FOR EACH ROW EXECUTE FUNCTION public.reroute_to_canonical();

CREATE OR REPLACE FUNCTION public.reroute_res_to_canonical()
RETURNS trigger LANGUAGE plpgsql SET search_path = public AS $fn$
DECLARE can uuid;
BEGIN
  SELECT canonical_item_id INTO can FROM public.inventory_items WHERE id = NEW.inventory_item_id;
  IF can IS NOT NULL AND can <> NEW.inventory_item_id THEN NEW.inventory_item_id := can; END IF;
  RETURN NEW;
END $fn$;

DROP TRIGGER IF EXISTS trg_agouza_res_reroute ON public.agouza_stock_reservations;
CREATE TRIGGER trg_agouza_res_reroute BEFORE INSERT ON public.agouza_stock_reservations
FOR EACH ROW EXECUTE FUNCTION public.reroute_res_to_canonical();

SAVEPOINT after_triggers;
""")

# ------------------------------------------------------------------
# Phase 8 — Assertions
# ------------------------------------------------------------------
add(f"""-- ------------------------------------------------------------------
-- Phase 8 — Assertions (rollback if any fail)
-- ------------------------------------------------------------------
DO $chk$
DECLARE
  kg_total numeric;
  fat_bal  numeric;
  suspended_bal numeric;
BEGIN
  SELECT COALESCE(SUM(balance),0) INTO kg_total
  FROM public.v_inventory_balances v
  JOIN public.inventory_items i ON i.id = v.canonical_item_id
  WHERE v.warehouse_id = '{MAIN_WH}'
    AND i.unit IN ('كيلو','كجم')
    AND i.canonical_item_id = i.id
    AND i.id <> '{FAT}'
    AND i.retired_at IS NULL;
  IF ROUND(kg_total::numeric,2) <> 624.50 THEN
    RAISE EXCEPTION 'Assertion failed: main WH kg total = %, expected 624.50', kg_total;
  END IF;

  SELECT COALESCE(balance,0) INTO fat_bal
  FROM public.v_inventory_balances
  WHERE canonical_item_id = '{FAT}' AND warehouse_id = '{MAIN_WH}';
  IF ROUND(fat_bal::numeric,2) <> 61.50 THEN
    RAISE EXCEPTION 'Assertion failed: fat pack balance = %, expected 61.50', fat_bal;
  END IF;
END $chk$;

-- Update main warehouse inventory_items.stock to reflect v_inventory_balances (kept in sync going forward by ledger)
UPDATE public.inventory_items i
SET stock = COALESCE(v.balance, 0)
FROM public.v_inventory_balances v
WHERE i.id = v.canonical_item_id
  AND i.warehouse_id = '{MAIN_WH}'
  AND i.canonical_item_id = i.id;

-- Retired aliases → stock = 0
UPDATE public.inventory_items SET stock = 0
WHERE warehouse_id = '{MAIN_WH}' AND retired_at IS NOT NULL;

INSERT INTO public.stocktake_20260727_checksums(label,row_count,qty_sum)
VALUES ('items_after', (SELECT count(*) FROM public.inventory_items WHERE warehouse_id='{MAIN_WH}'),
        (SELECT SUM(stock) FROM public.inventory_items WHERE warehouse_id='{MAIN_WH}' AND retired_at IS NULL));

COMMIT;
-- ============================================================================
-- End of forward migration {STK}
-- ============================================================================
""")

with open(FWD,'w') as f: f.write('\n'.join(out))
print('WROTE', FWD, len(out), 'lines')

# ---------- ROLLBACK ----------
r = []
r.append(f"""-- ============================================================================
-- Rollback: {STK}
-- Restores main-warehouse rows to their pre-migration state using the snapshot
-- tables created by the forward migration. NO DELETE against production rows.
-- Requires GM/Executive role.
-- ============================================================================
BEGIN;
SET LOCAL statement_timeout = '10min';

DO $g$
DECLARE r text; ok boolean := false;
BEGIN
  FOR r IN SELECT role::text FROM public.user_roles WHERE user_id = auth.uid()
  LOOP IF r IN ('general_manager','executive_manager') THEN ok := true; END IF; END LOOP;
  IF NOT ok THEN RAISE EXCEPTION 'Only GM/Executive can rollback {STK}'; END IF;
END $g$;

-- 1) Restore inventory_movements rows for the main warehouse from snapshot (actual pre-values).
--    All movements inserted by the migration carry stocktake_ref='{STK}' → remove them.
DELETE FROM public.inventory_movements
WHERE warehouse_id = '{MAIN_WH}'
  AND stocktake_ref = '{STK}'
  AND idempotency_key IS NOT NULL;

INSERT INTO public.inventory_movements AS m
SELECT * FROM public.inventory_movements_snapshot_20260727 s
ON CONFLICT (id) DO UPDATE SET
  item_id                 = EXCLUDED.item_id,
  warehouse_id            = EXCLUDED.warehouse_id,
  movement_type           = EXCLUDED.movement_type,
  quantity                = EXCLUDED.quantity,
  destination_warehouse_id= EXCLUDED.destination_warehouse_id,
  reference               = EXCLUDED.reference,
  party                   = EXCLUDED.party,
  unit_cost               = EXCLUDED.unit_cost,
  notes                   = EXCLUDED.notes,
  performed_by            = EXCLUDED.performed_by,
  performed_at            = EXCLUDED.performed_at,
  movement_no             = EXCLUDED.movement_no,
  module                  = EXCLUDED.module,
  source_warehouse_id     = EXCLUDED.source_warehouse_id,
  reference_type          = EXCLUDED.reference_type,
  reference_id            = EXCLUDED.reference_id,
  batch_id                = EXCLUDED.batch_id,
  reason                  = EXCLUDED.reason,
  approval_status         = EXCLUDED.approval_status,
  approved_by             = EXCLUDED.approved_by,
  approved_at             = EXCLUDED.approved_at,
  total_cost              = EXCLUDED.total_cost,
  order_item_id           = EXCLUDED.order_item_id,
  product_id              = EXCLUDED.product_id,
  package_count           = EXCLUDED.package_count,
  package_weight_kg       = EXCLUDED.package_weight_kg,
  quantity_kg             = EXCLUDED.quantity_kg;

-- Wipe reversal metadata for rows the migration only status-flipped
UPDATE public.inventory_movements
SET reversal_of_movement_id = NULL,
    reversed_at             = NULL,
    reversed_by             = NULL,
    reversal_reason         = NULL,
    stocktake_ref           = NULL
WHERE stocktake_ref = '{STK}';

-- 2) Restore inventory_items rows for the main warehouse
INSERT INTO public.inventory_items AS i
SELECT * FROM public.inventory_items_snapshot_20260727 s
ON CONFLICT (id) DO UPDATE SET
  warehouse_id        = EXCLUDED.warehouse_id,
  name                = EXCLUDED.name,
  category            = EXCLUDED.category,
  sku                 = EXCLUDED.sku,
  unit                = EXCLUDED.unit,
  stock               = EXCLUDED.stock,
  low_stock_threshold = EXCLUDED.low_stock_threshold,
  unit_cost           = EXCLUDED.unit_cost,
  expiry_date         = EXCLUDED.expiry_date,
  notes               = EXCLUDED.notes,
  is_active           = EXCLUDED.is_active,
  updated_at          = EXCLUDED.updated_at,
  reserved_qty        = EXCLUDED.reserved_qty,
  blocked_qty         = EXCLUDED.blocked_qty,
  module              = EXCLUDED.module,
  item_code           = EXCLUDED.item_code,
  last_movement_date  = EXCLUDED.last_movement_date,
  product_id          = EXCLUDED.product_id;

UPDATE public.inventory_items
SET canonical_item_id = NULL,
    retired_at        = NULL,
    retired_reason    = NULL
WHERE warehouse_id = '{MAIN_WH}';

-- 3) Drop schema additions created by the forward migration
DROP TRIGGER IF EXISTS trg_inv_mov_reroute_canonical ON public.inventory_movements;
DROP TRIGGER IF EXISTS trg_agouza_res_reroute       ON public.agouza_stock_reservations;
DROP FUNCTION IF EXISTS public.reroute_to_canonical();
DROP FUNCTION IF EXISTS public.reroute_res_to_canonical();
DROP VIEW      IF EXISTS public.v_inventory_balances;
DROP INDEX     IF EXISTS public.ux_inv_mov_idempotency_key;

ALTER TABLE public.inventory_movements
  DROP COLUMN IF EXISTS idempotency_key,
  DROP COLUMN IF EXISTS reversal_of_movement_id,
  DROP COLUMN IF EXISTS reversed_at,
  DROP COLUMN IF EXISTS reversed_by,
  DROP COLUMN IF EXISTS reversal_reason,
  DROP COLUMN IF EXISTS stocktake_ref;

ALTER TABLE public.inventory_items
  DROP COLUMN IF EXISTS canonical_item_id,
  DROP COLUMN IF EXISTS retired_at,
  DROP COLUMN IF EXISTS retired_reason;

-- 4) Verify row counts against snapshot checksums
DO $v$
DECLARE want bigint; got bigint;
BEGIN
  SELECT row_count INTO want FROM public.stocktake_20260727_checksums WHERE label='items_before';
  SELECT count(*) INTO got  FROM public.inventory_items WHERE warehouse_id='{MAIN_WH}';
  IF want <> got THEN RAISE EXCEPTION 'Rollback: items count mismatch want=% got=%', want, got; END IF;
END $v$;

COMMIT;
-- After confirming production is healthy, snapshot tables may be dropped:
--   DROP TABLE public.inventory_items_snapshot_20260727;
--   DROP TABLE public.inventory_movements_snapshot_20260727;
--   DROP TABLE public.stocktake_20260727_checksums;
""")
with open(RBK,'w') as f: f.write('\n'.join(r))
print('WROTE', RBK)
