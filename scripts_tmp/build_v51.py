"""Build v5.1 package: fixed xlsx (summary/product/6ك), forward SQL (126 movements, view/security/locks/actor/lines), rollback SQL (explicit columns)."""
import openpyxl, hashlib
from copy import copy
from datetime import datetime, timezone

V5   = '/mnt/documents/DryRun_MAIN-STOCKTAKE-20260727_v5.xlsx'
V51  = '/mnt/documents/DryRun_MAIN-STOCKTAKE-20260727_v5_1.xlsx'
FWD  = '/mnt/documents/forward_migration_MAIN-STOCKTAKE-20260727_v5_1.sql'
RBK  = '/mnt/documents/rollback_MAIN-STOCKTAKE-20260727_v5_1.sql'
STK      = 'MAIN-STOCKTAKE-20260727'
MAIN_WH  = '5ec781b5-685b-4806-b59a-83a79ea5662c'
FAT      = 'a55763c7-6e12-4ed0-8d62-37ace6bf86a3'   # دهن النعام (عبوة) — independent
PIN_6K   = '8c304eb4-286a-43db-9d2e-ff6bde3e17d7'   # placeholder; overwritten below by name lookup

wb = openpyxl.load_workbook(V5)

def rows(sheet):
    ws = wb[sheet]
    hdr = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None and all(v is None for v in row): continue
        yield dict(zip(hdr, row))

# ---- discover 6ك canonical from Item Master ----
for r in rows('Item Master Action Plan'):
    nm = (r.get('Final Name') or '') + ' ' + (r.get('Current Name') or '')
    if '6ك' in nm and 'دبوس' in nm and 'عظم' in nm:
        PIN_6K = r['Canonical Item ID']
        break

def q(v):
    if v is None: return 'NULL'
    if isinstance(v, bool): return 'true' if v else 'false'
    if isinstance(v, (int, float)): return repr(v) if isinstance(v,float) else str(v)
    return "'" + str(v).replace("'", "''") + "'"

# =========================================================================
# XLSX v5.1: update Summary + Product Link Summary + Alias Retirement Plan
# =========================================================================
# Count product link actions
plan_rewire = plan_keep = plan_independent = 0
for r in rows('Product Link Action Plan'):
    src = r['Current inventory_item_id (source)']
    can = r['Canonical Item ID']
    if src == FAT or can == FAT:
        plan_independent += 1
    elif src == can:
        plan_keep += 1
    else:
        plan_rewire += 1

# Fix Summary sheet
if 'Summary' in wb.sheetnames:
    ws = wb['Summary']
    # append clarifying rows at bottom
    ws.append([])
    ws.append(['--- v5.1 CORRECTIONS ---'])
    ws.append(['Movement rows in plan', 126, '100 reclassification legs + 26 stocktake adjustments'])
    ws.append(['SQL INSERT statements (Phase 5+6)', 126, 'one INSERT per row in Movement Action Plan (no self-pairing)'])
    ws.append(['Product Link Actions', 27, f'{plan_rewire} rewire + {plan_keep} keep + {plan_independent} independent (fat pack)'])
    ws.append(['6ك دبوس بالعظم canonical', PIN_6K, 'Active; final balance = 0 via posted STOCKTAKE_ADJUSTMENT (documented)'])
    ws.append(['دهن النعام (عبوة)', FAT, 'independent; not aliased; 61.5 packs; excluded from 624.5 kg total'])
    ws.append(['View v_inventory_balances', 'security_invoker=true, GROUP BY canonical_item_id only (1 row per canonical)'])
    ws.append(['Snapshot tables', 'REVOKE ALL from PUBLIC/anon/authenticated; access via service_role only'])

# Save v5.1 workbook
wb.save(V51)
print('WROTE', V51)

# =========================================================================
# FORWARD MIGRATION SQL v5.1
# =========================================================================
now_iso = datetime.now(timezone.utc).isoformat()

# Full inventory_movements column list (explicit — used in snapshot restore)
IM_COLS = ['id','item_id','warehouse_id','movement_type','quantity','destination_warehouse_id',
           'reference','party','unit_cost','notes','performed_by','performed_at','created_at',
           'movement_no','module','source_warehouse_id','reference_type','reference_id',
           'batch_id','reason','approval_status','approved_by','approved_at','total_cost',
           'order_item_id','product_id','package_count','package_weight_kg','quantity_kg']
II_COLS = ['id','warehouse_id','name','category','sku','unit','stock','low_stock_threshold',
           'unit_cost','expiry_date','notes','is_active','created_at','updated_at',
           'reserved_qty','blocked_qty','module','item_code','last_movement_date','product_id']

out = []
add = out.append

add(f"""-- ============================================================================
-- MAIN-STOCKTAKE-20260727 — Forward Migration v5.1
-- Scope: warehouse_id = {MAIN_WH} ONLY (المخزن الرئيسي - المقر)
-- Generated: {now_iso}
-- Correction summary vs v5:
--   1. Emits ONE INSERT per Movement Action Plan row (126 total, not 226).
--   2. v_inventory_balances GROUP BY canonical_item_id, warehouse_id only
--      (one row per canonical) + WITH (security_invoker = true).
--   3. Snapshot tables get REVOKE ALL from PUBLIC/anon/authenticated.
--   4. LOCK TABLE on the two hot tables inside the transaction.
--   5. Actor UUID resolved from :actor_id psql var when auth.uid() is NULL.
--   6. reroute_to_canonical trigger SKIPS rows with stocktake_ref = STK
--      so alias→canonical OUT legs actually post to the alias.
--   7. Preflight assertions: schema objects & idempotency keys must not exist.
--   8. Independent assertion for 6ك دبوس بالعظم canonical = 0.
--   9. Correct stocktaking_sessions columns + 27 stocktaking_lines rows.
--  10. Explicit column lists (no SELECT * anywhere in the reversible path).
-- Safety: single BEGIN/COMMIT + SAVEPOINTs; DO NOT RUN without written approval.
-- Required psql var:  \\set actor_id '<uuid-of-executing-admin>'
-- ============================================================================
\\set ON_ERROR_STOP on
\\if :{{?actor_id}}
\\else
  \\echo 'ERROR: psql variable :actor_id must be set (executor UUID).'
  \\quit 1
\\endif

BEGIN;
SET LOCAL statement_timeout = '15min';
SET LOCAL lock_timeout      = '30s';
SET LOCAL idle_in_transaction_session_timeout = '10min';

-- ------------------------------------------------------------------
-- Phase -1 — Preflight (actor + schema-object non-existence)
-- ------------------------------------------------------------------
DO $pre$
DECLARE
  v_actor uuid;
  v_role_ok boolean;
BEGIN
  v_actor := COALESCE(auth.uid(), :'actor_id'::uuid);
  IF v_actor IS NULL THEN
    RAISE EXCEPTION 'Preflight: no actor available (auth.uid() NULL and :actor_id not provided)';
  END IF;
  PERFORM set_config('app.stk_actor', v_actor::text, true);

  SELECT bool_or(role::text IN ('general_manager','executive_manager'))
    INTO v_role_ok
  FROM public.user_roles WHERE user_id = v_actor;
  IF NOT COALESCE(v_role_ok,false) THEN
    RAISE EXCEPTION 'Preflight: actor % lacks GM/Executive role', v_actor;
  END IF;

  -- Schema objects introduced by this migration must not already exist
  IF EXISTS (SELECT 1 FROM information_schema.columns
             WHERE table_schema='public' AND table_name='inventory_items'
               AND column_name='canonical_item_id') THEN
    RAISE EXCEPTION 'Preflight: inventory_items.canonical_item_id already exists (prior run?)';
  END IF;
  IF EXISTS (SELECT 1 FROM information_schema.columns
             WHERE table_schema='public' AND table_name='inventory_movements'
               AND column_name='idempotency_key') THEN
    RAISE EXCEPTION 'Preflight: inventory_movements.idempotency_key already exists (prior run?)';
  END IF;
  IF to_regclass('public.inventory_items_snapshot_20260727') IS NOT NULL THEN
    RAISE EXCEPTION 'Preflight: snapshot table already exists (prior run?)';
  END IF;
  IF to_regclass('public.v_inventory_balances') IS NOT NULL THEN
    RAISE EXCEPTION 'Preflight: v_inventory_balances view already exists (prior run?)';
  END IF;
END $pre$;

-- ------------------------------------------------------------------
-- Phase 0 — Lock and snapshot main-warehouse rows
-- ------------------------------------------------------------------
LOCK TABLE public.inventory_items       IN SHARE ROW EXCLUSIVE MODE;
LOCK TABLE public.inventory_movements   IN SHARE ROW EXCLUSIVE MODE;

CREATE TABLE public.inventory_items_snapshot_20260727
  (LIKE public.inventory_items INCLUDING ALL);
INSERT INTO public.inventory_items_snapshot_20260727 ({ii_cols})
SELECT {ii_cols} FROM public.inventory_items WHERE warehouse_id = '{MAIN_WH}';

CREATE TABLE public.inventory_movements_snapshot_20260727
  (LIKE public.inventory_movements INCLUDING ALL);
INSERT INTO public.inventory_movements_snapshot_20260727 ({im_cols})
SELECT {im_cols} FROM public.inventory_movements WHERE warehouse_id = '{MAIN_WH}';

-- Lock down snapshots (never expose via API)
REVOKE ALL ON public.inventory_items_snapshot_20260727     FROM PUBLIC, anon, authenticated;
REVOKE ALL ON public.inventory_movements_snapshot_20260727 FROM PUBLIC, anon, authenticated;
GRANT  ALL ON public.inventory_items_snapshot_20260727     TO service_role;
GRANT  ALL ON public.inventory_movements_snapshot_20260727 TO service_role;

CREATE TABLE public.stocktake_20260727_checksums (
  label text PRIMARY KEY,
  row_count bigint NOT NULL,
  qty_sum numeric,
  captured_at timestamptz NOT NULL DEFAULT now()
);
REVOKE ALL ON public.stocktake_20260727_checksums FROM PUBLIC, anon, authenticated;
GRANT  ALL ON public.stocktake_20260727_checksums TO service_role;

INSERT INTO public.stocktake_20260727_checksums(label,row_count,qty_sum)
SELECT 'items_before_main', count(*)::bigint, sum(stock)
FROM public.inventory_items_snapshot_20260727;
INSERT INTO public.stocktake_20260727_checksums(label,row_count,qty_sum)
SELECT 'movements_before_main', count(*)::bigint, NULL
FROM public.inventory_movements_snapshot_20260727;
INSERT INTO public.stocktake_20260727_checksums(label,row_count,qty_sum)
SELECT 'items_before_other_wh', count(*)::bigint, sum(stock)
FROM public.inventory_items WHERE warehouse_id <> '{MAIN_WH}';

SAVEPOINT after_snapshot;

-- ------------------------------------------------------------------
-- Phase 1 — Schema additions
-- ------------------------------------------------------------------
ALTER TABLE public.inventory_items
  ADD COLUMN canonical_item_id uuid REFERENCES public.inventory_items(id),
  ADD COLUMN retired_at        timestamptz,
  ADD COLUMN retired_reason    text;

ALTER TABLE public.inventory_movements
  ADD COLUMN idempotency_key         text,
  ADD COLUMN reversal_of_movement_id uuid REFERENCES public.inventory_movements(id),
  ADD COLUMN reversed_at             timestamptz,
  ADD COLUMN reversed_by             uuid REFERENCES auth.users(id),
  ADD COLUMN reversal_reason         text,
  ADD COLUMN stocktake_ref           text;

CREATE UNIQUE INDEX ux_inv_mov_idempotency_key
  ON public.inventory_movements(idempotency_key)
  WHERE idempotency_key IS NOT NULL;

SAVEPOINT after_schema;
""".format(ii_cols=', '.join(II_COLS), im_cols=', '.join(IM_COLS)))

# ---------- Phase 2 — Item Master (27 canonicals) ----------
add("-- ------------------------------------------------------------------")
add("-- Phase 2 — Item Master: 27 canonicals updated (name/unit/barcode/self-canonical)")
add("-- ------------------------------------------------------------------")
n_can = 0
for r in rows('Item Master Action Plan'):
    cid   = r['Canonical Item ID']
    fname = r['Final Name']
    funit = r['Final Unit']
    fbc   = r['Final Barcode (text)']
    add(f"UPDATE public.inventory_items SET "
        f"name={q(fname)}, unit={q(funit)}, "
        f"sku=COALESCE(NULLIF(sku,''), {q(fbc)}), "
        f"canonical_item_id=id, is_active=true, retired_at=NULL, retired_reason=NULL "
        f"WHERE id={q(cid)} AND warehouse_id={q(MAIN_WH)};")
    n_can += 1
add(f"DO $c$ BEGIN IF (SELECT count(*) FROM public.inventory_items "
    f"WHERE warehouse_id={q(MAIN_WH)} AND canonical_item_id = id AND retired_at IS NULL) < {n_can} "
    f"THEN RAISE EXCEPTION 'Phase 2: canonicals updated < expected {n_can}'; END IF; END $c$;")
add("SAVEPOINT after_canonicals;\n")

# ---------- Phase 3 — Alias retirement ----------
add("-- ------------------------------------------------------------------")
add("-- Phase 3 — Alias retirement: canonical_item_id + is_active=false + retired_at")
add("-- ------------------------------------------------------------------")
n_al = 0
for r in rows('Alias Retirement Plan'):
    sid = r['Source Item ID']
    cid = r['Canonical Item ID']
    if sid == FAT: continue
    add(f"UPDATE public.inventory_items SET "
        f"canonical_item_id={q(cid)}, is_active=false, "
        f"retired_at=now(), retired_reason={q(STK)} "
        f"WHERE id={q(sid)} AND warehouse_id={q(MAIN_WH)};")
    n_al += 1
add(f"-- {n_al} alias rows retired.")
add("SAVEPOINT after_aliases;\n")

# ---------- Phase 4 — Product link rewires ----------
add("-- ------------------------------------------------------------------")
add(f"-- Phase 4 — Product Link Actions: {plan_rewire} rewire + {plan_keep} keep + {plan_independent} independent (fat)")
add("-- ------------------------------------------------------------------")
n_rw = 0
for r in rows('Product Link Action Plan'):
    pid = r['Product ID']
    src = r['Current inventory_item_id (source)']
    can = r['Canonical Item ID']
    if src == FAT or can == FAT: continue
    if src == can: continue
    add(f"-- rewire product {pid}: {src} -> {can}")
    add(f"UPDATE public.inventory_items SET product_id=NULL "
        f"WHERE id={q(src)} AND warehouse_id={q(MAIN_WH)} AND product_id={q(pid)};")
    add(f"UPDATE public.inventory_items SET product_id={q(pid)} "
        f"WHERE id={q(can)} AND warehouse_id={q(MAIN_WH)} "
        f"AND (product_id IS NULL OR product_id={q(pid)});")
    n_rw += 1
add(f"-- {n_rw} product link rewires applied.")
add("SAVEPOINT after_product_links;\n")

# ---------- Phase 5 — 104 opening reversals (status flip) ----------
add("-- ------------------------------------------------------------------")
add("-- Phase 5 — 104 duplicate opening reversals (status flip; NO new movement)")
add("-- ------------------------------------------------------------------")
n_rev = 0
for r in rows('Opening Reversal Actions'):
    mid  = r['Movement ID (to flip)']
    kept = r['Kept Opening Movement ID']
    add(f"UPDATE public.inventory_movements SET "
        f"approval_status='reversed', reversed_at=now(), "
        f"reversed_by=current_setting('app.stk_actor')::uuid, "
        f"reversal_of_movement_id={q(kept)}, "
        f"reversal_reason={q('duplicate opening; kept=' + str(kept))}, "
        f"stocktake_ref={q(STK)} "
        f"WHERE id={q(mid)} AND warehouse_id={q(MAIN_WH)} AND approval_status='posted';")
    n_rev += 1
add(f"DO $r$ BEGIN IF (SELECT count(*) FROM public.inventory_movements "
    f"WHERE warehouse_id={q(MAIN_WH)} AND stocktake_ref={q(STK)} "
    f"AND approval_status='reversed') <> {n_rev} "
    f"THEN RAISE EXCEPTION 'Phase 5: expected {n_rev} reversed openings'; END IF; END $r$;")
add("SAVEPOINT after_openings;\n")

# ---------- Phase 6 — Reroute trigger (must SKIP stocktake movements) ----------
add("""-- ------------------------------------------------------------------
-- Phase 6a — Reroute trigger (skips rows carrying stocktake_ref)
-- ------------------------------------------------------------------
CREATE OR REPLACE FUNCTION public.reroute_to_canonical()
RETURNS trigger LANGUAGE plpgsql SET search_path = public AS $fn$
DECLARE can uuid;
BEGIN
  -- Stocktake-owned rows must land on the item_id they name (alias legs must post to the alias)
  IF NEW.stocktake_ref IS NOT NULL THEN
    RETURN NEW;
  END IF;
  SELECT canonical_item_id INTO can FROM public.inventory_items WHERE id = NEW.item_id;
  IF can IS NOT NULL AND can <> NEW.item_id THEN NEW.item_id := can; END IF;
  RETURN NEW;
END $fn$;

DROP TRIGGER IF EXISTS trg_inv_mov_reroute_canonical ON public.inventory_movements;
CREATE TRIGGER trg_inv_mov_reroute_canonical BEFORE INSERT ON public.inventory_movements
FOR EACH ROW EXECUTE FUNCTION public.reroute_to_canonical();

CREATE OR REPLACE FUNCTION public.reroute_res_to_canonical()
RETURNS trigger LANGUAGE plpgsql SET search_path = public AS $fn$
DECLARE can uuid;
BEGIN
  SELECT canonical_item_id INTO can FROM public.inventory_items WHERE id = NEW.inventory_item_id;
  IF can IS NOT NULL AND can <> NEW.inventory_item_id THEN NEW.inventory_item_id := can; END IF;
  RETURN NEW;
END $fn$;

DROP TRIGGER IF EXISTS trg_agouza_res_reroute ON public.agouza_stock_reservations;
CREATE TRIGGER trg_agouza_res_reroute BEFORE INSERT ON public.agouza_stock_reservations
FOR EACH ROW EXECUTE FUNCTION public.reroute_res_to_canonical();
SAVEPOINT after_triggers;
""")

# ---------- Phase 6b — 126 movements (one per Movement Action Plan row) ----------
add("-- ------------------------------------------------------------------")
add("-- Phase 6b — 126 posted movements (one per Movement Action Plan row)")
add("--            100 reclassification legs (50 OUT + 50 IN) + 26 adjustments")
add("-- ------------------------------------------------------------------")
n_reclass = n_adjust = 0
for r in rows('Movement Action Plan'):
    op   = r['Operation Type']
    sid  = r['Source Item ID']
    cid  = r['Canonical Item ID']
    qty  = r['Quantity']
    direction = str(r['Direction'] or '')
    idem = r['Idempotency Key']
    ref_type = r['Reference Type']
    notes = r['notes']
    mt = r['movement_type']
    # Determine target item_id and DB movement_type ('in' or 'out')
    if op == 'RECLASSIFY':
        # OUT (alias→canonical) posts on alias; IN posts on canonical
        item_id  = sid if direction.startswith('OUT') else cid
        db_mt    = 'out' if direction.startswith('OUT') else 'in'
        n_reclass += 1
    elif op == 'RECLASSIFY_NEGATIVE':
        # Repair: bring canonical toward zero and hold on alias (or vice versa per plan direction)
        item_id  = sid if direction.startswith('OUT') else cid
        db_mt    = 'out' if direction.startswith('OUT') else 'in'
        n_reclass += 1
    else:  # STOCKTAKE_ADJUSTMENT
        item_id  = cid
        db_mt    = 'out' if direction.startswith('OUT') else 'in'
        n_adjust += 1
    add(f"-- {op} {direction} qty={qty} on {item_id}  ({notes})")
    add(f"INSERT INTO public.inventory_movements "
        f"(item_id, warehouse_id, movement_type, quantity, unit_cost, notes, "
        f" reference_type, reference_id, stocktake_ref, idempotency_key, "
        f" approval_status, performed_by, approved_by, approved_at) "
        f"VALUES ({q(item_id)}, {q(MAIN_WH)}, {q(db_mt)}, {q(qty)}, 0, {q(notes)}, "
        f"       {q(ref_type)}, {q(STK)}, {q(STK)}, {q(idem)}, "
        f"       'posted', current_setting('app.stk_actor')::uuid, "
        f"       current_setting('app.stk_actor')::uuid, now());")

add(f"-- reclassification legs inserted: {n_reclass} (expected 100)")
add(f"-- stocktake adjustments inserted: {n_adjust} (expected 26)")
add(f"DO $m$ DECLARE c int; BEGIN "
    f"SELECT count(*) INTO c FROM public.inventory_movements "
    f"WHERE warehouse_id={q(MAIN_WH)} AND stocktake_ref={q(STK)} AND idempotency_key IS NOT NULL; "
    f"IF c <> {n_reclass + n_adjust} THEN "
    f"  RAISE EXCEPTION 'Phase 6b: inserted % movements, expected {n_reclass + n_adjust}', c; "
    f"END IF; END $m$;")
add("SAVEPOINT after_movements;\n")

# ---------- Phase 7 — Stocktake session + lines ----------
add(f"""-- ------------------------------------------------------------------
-- Phase 7 — Stocktake session header + 27 lines (one per canonical)
-- ------------------------------------------------------------------
WITH s AS (
  INSERT INTO public.stocktaking_sessions
    (session_no, warehouse_id, count_date, stocktaker_name, status, notes, created_by, approved_by, approved_at, reference_id)
  VALUES ({q(STK)}, {q(MAIN_WH)}, DATE '2026-07-27',
          'Physical Count Team', 'approved',
          'MAIN-STOCKTAKE-20260727: 26 kg items totalling 624.5 kg + دبوس 6ك بالعظم=0',
          current_setting('app.stk_actor')::uuid,
          current_setting('app.stk_actor')::uuid, now(), {q(STK)})
  RETURNING id
)
INSERT INTO public.stocktaking_lines (session_id, item_id, system_qty, actual_qty, unit_cost, reason, notes)
SELECT s.id, v.item_id, v.system_qty, v.actual_qty, 0, v.reason, v.notes
FROM s, (VALUES
""")

# Build the VALUES: 27 canonicals with actual quantities from Adjustments sheet if available
lines = []
adj_actual = {}
if 'Adjustments' in wb.sheetnames:
    for r in rows('Adjustments'):
        cid = r.get('Canonical Item ID') or r.get('Item ID')
        if cid:
            adj_actual[cid] = r.get('Physical Count') or r.get('Actual Qty') or r.get('actual_qty') or 0
# Fallback: derive from Movement Action Plan STOCKTAKE_ADJUSTMENT notes ("physical count = N")
for r in rows('Movement Action Plan'):
    if r['Operation Type'] == 'STOCKTAKE_ADJUSTMENT':
        n = str(r['notes'] or '')
        cid = r['Canonical Item ID']
        if 'physical count' in n:
            try:
                adj_actual.setdefault(cid, float(n.split('=')[-1].strip()))
            except: pass

for r in rows('Item Master Action Plan'):
    cid = r['Canonical Item ID']
    fname = r['Final Name'] or ''
    actual = adj_actual.get(cid, 0)
    is_pin6k = (cid == PIN_6K)
    if is_pin6k: actual = 0
    reason = 'stocktake_MAIN-STOCKTAKE-20260727'
    note = fname
    lines.append(f"({q(cid)}::uuid, 0, {q(actual)}, {q(reason)}, {q(note)})")
add(",\n".join(lines))
add(") AS v(item_id, system_qty, actual_qty, reason, notes);")
add("SAVEPOINT after_session;\n")

# ---------- Phase 8 — v_inventory_balances (canonical-grouped, security_invoker) ----------
add(f"""-- ------------------------------------------------------------------
-- Phase 8 — Balance view: ONE row per (canonical_item_id, warehouse_id)
-- ------------------------------------------------------------------
CREATE VIEW public.v_inventory_balances
WITH (security_invoker = true) AS
SELECT
  COALESCE(i.canonical_item_id, m.item_id) AS canonical_item_id,
  m.warehouse_id,
  SUM(CASE
        WHEN m.movement_type IN ('out','sales_dispatch','stock_out',
                                 'production_consumption','packaging_consumption','waste_loss')
          THEN -m.quantity ELSE m.quantity END) AS balance
FROM public.inventory_movements m
LEFT JOIN public.inventory_items i ON i.id = m.item_id
WHERE m.approval_status = 'posted'
GROUP BY 1, 2;
GRANT SELECT ON public.v_inventory_balances TO authenticated, service_role;

-- Detail view (per raw item) — for audit only, NOT used for stock sync
CREATE VIEW public.v_inventory_balances_raw
WITH (security_invoker = true) AS
SELECT
  m.item_id AS raw_item_id,
  COALESCE(i.canonical_item_id, m.item_id) AS canonical_item_id,
  m.warehouse_id,
  SUM(CASE
        WHEN m.movement_type IN ('out','sales_dispatch','stock_out',
                                 'production_consumption','packaging_consumption','waste_loss')
          THEN -m.quantity ELSE m.quantity END) AS balance
FROM public.inventory_movements m
LEFT JOIN public.inventory_items i ON i.id = m.item_id
WHERE m.approval_status = 'posted'
GROUP BY 1, 2, 3;
GRANT SELECT ON public.v_inventory_balances_raw TO authenticated, service_role;
SAVEPOINT after_view;
""")

# ---------- Phase 9 — Stock sync + Assertions ----------
add(f"""-- ------------------------------------------------------------------
-- Phase 9 — Sync inventory_items.stock from canonical-grouped view
-- ------------------------------------------------------------------
UPDATE public.inventory_items i
SET stock = COALESCE(v.balance, 0)
FROM (
  SELECT canonical_item_id, warehouse_id, SUM(balance) AS balance
  FROM public.v_inventory_balances
  WHERE warehouse_id = '{MAIN_WH}'
  GROUP BY canonical_item_id, warehouse_id
) v
WHERE i.id = v.canonical_item_id
  AND i.warehouse_id = '{MAIN_WH}'
  AND i.canonical_item_id = i.id;

-- Retired aliases in main WH: force stock=0
UPDATE public.inventory_items SET stock = 0
WHERE warehouse_id = '{MAIN_WH}' AND retired_at IS NOT NULL;

-- ------------------------------------------------------------------
-- Phase 10 — Assertions (any failure aborts the transaction)
-- ------------------------------------------------------------------
DO $chk$
DECLARE
  kg_total       numeric;
  fat_bal        numeric;
  pin6k_bal      numeric;
  dup_canonical  int;
  other_wh_count bigint;
  snap_other     bigint;
BEGIN
  -- 10.1 kg subtotal (main WH, canonicals only, exclude fat pack, exclude retired)
  SELECT COALESCE(SUM(v.balance),0) INTO kg_total
  FROM public.v_inventory_balances v
  JOIN public.inventory_items i ON i.id = v.canonical_item_id
  WHERE v.warehouse_id = '{MAIN_WH}'
    AND i.unit IN ('كيلو','كجم')
    AND i.canonical_item_id = i.id
    AND i.id <> '{FAT}'
    AND i.retired_at IS NULL;
  IF ROUND(kg_total::numeric,2) <> 624.50 THEN
    RAISE EXCEPTION 'Assertion 10.1: kg total = %, expected 624.50', kg_total;
  END IF;

  -- 10.2 Fat pack independent
  SELECT COALESCE(SUM(balance),0) INTO fat_bal
  FROM public.v_inventory_balances
  WHERE canonical_item_id = '{FAT}' AND warehouse_id = '{MAIN_WH}';
  IF ROUND(fat_bal::numeric,2) <> 61.50 THEN
    RAISE EXCEPTION 'Assertion 10.2: fat pack balance = %, expected 61.50', fat_bal;
  END IF;

  -- 10.3 دبوس 6ك بالعظم canonical = 0 (independent, not covered by 10.1 alone)
  SELECT COALESCE(SUM(balance),0) INTO pin6k_bal
  FROM public.v_inventory_balances
  WHERE canonical_item_id = '{PIN_6K}' AND warehouse_id = '{MAIN_WH}';
  IF ROUND(pin6k_bal::numeric,4) <> 0 THEN
    RAISE EXCEPTION 'Assertion 10.3: 6ك دبوس بالعظم balance = %, expected 0', pin6k_bal;
  END IF;

  -- 10.4 View returns exactly one row per canonical in main WH
  SELECT count(*) INTO dup_canonical FROM (
    SELECT canonical_item_id FROM public.v_inventory_balances
    WHERE warehouse_id = '{MAIN_WH}'
    GROUP BY canonical_item_id HAVING count(*) > 1
  ) d;
  IF dup_canonical <> 0 THEN
    RAISE EXCEPTION 'Assertion 10.4: v_inventory_balances has % duplicated canonicals in main WH', dup_canonical;
  END IF;

  -- 10.5 Other-warehouse item count unchanged
  SELECT count(*) INTO other_wh_count FROM public.inventory_items WHERE warehouse_id <> '{MAIN_WH}';
  SELECT row_count INTO snap_other FROM public.stocktake_20260727_checksums WHERE label='items_before_other_wh';
  IF other_wh_count <> snap_other THEN
    RAISE EXCEPTION 'Assertion 10.5: other-warehouse item count changed (% vs %)', other_wh_count, snap_other;
  END IF;
END $chk$;

INSERT INTO public.stocktake_20260727_checksums(label,row_count,qty_sum)
SELECT 'items_after_main', count(*)::bigint, sum(stock)
FROM public.inventory_items WHERE warehouse_id='{MAIN_WH}';
INSERT INTO public.stocktake_20260727_checksums(label,row_count,qty_sum)
SELECT 'movements_after_main', count(*)::bigint, NULL
FROM public.inventory_movements WHERE warehouse_id='{MAIN_WH}';

COMMIT;
-- ============================================================================
-- End of MAIN-STOCKTAKE-20260727 forward migration v5.1
-- ============================================================================
""")

with open(FWD,'w') as f: f.write('\n'.join(out))
print('WROTE', FWD, len(out), 'lines')

# =========================================================================
# ROLLBACK SQL v5.1  (explicit columns; no SELECT *)
# =========================================================================
r = []
r.append(f"""-- ============================================================================
-- MAIN-STOCKTAKE-20260727 — Rollback v5.1
-- Restores main-warehouse rows to their pre-migration state using explicit
-- column lists that match the snapshot schema (safe against Phase-1 ADD COLUMN).
-- Requires:  \\set actor_id '<uuid-of-executing-admin>'
-- ============================================================================
\\set ON_ERROR_STOP on
\\if :{{?actor_id}}
\\else
  \\echo 'ERROR: :actor_id required.'
  \\quit 1
\\endif

BEGIN;
SET LOCAL statement_timeout = '15min';
SET LOCAL lock_timeout      = '30s';

DO $g$
DECLARE v_actor uuid; ok boolean := false;
BEGIN
  v_actor := COALESCE(auth.uid(), :'actor_id'::uuid);
  IF v_actor IS NULL THEN RAISE EXCEPTION 'Rollback: no actor'; END IF;
  SELECT bool_or(role::text IN ('general_manager','executive_manager')) INTO ok
  FROM public.user_roles WHERE user_id = v_actor;
  IF NOT COALESCE(ok,false) THEN RAISE EXCEPTION 'Rollback: actor lacks GM/Executive role'; END IF;
END $g$;

LOCK TABLE public.inventory_items     IN SHARE ROW EXCLUSIVE MODE;
LOCK TABLE public.inventory_movements IN SHARE ROW EXCLUSIVE MODE;

-- 1) Remove movements INSERTED by the migration (they all carry idempotency_key + stocktake_ref)
DELETE FROM public.inventory_movements
WHERE warehouse_id = '{MAIN_WH}'
  AND stocktake_ref = '{STK}'
  AND idempotency_key IS NOT NULL;

-- 2) Restore inventory_movements rows (status flips + any updated columns) from snapshot
INSERT INTO public.inventory_movements AS m ({im_cols})
SELECT {im_cols} FROM public.inventory_movements_snapshot_20260727 s
ON CONFLICT (id) DO UPDATE SET
""".format(im_cols=', '.join(IM_COLS)))
# Non-id columns updated on conflict
set_cols = [c for c in IM_COLS if c != 'id']
r.append(',\n'.join(f'  {c} = EXCLUDED.{c}' for c in set_cols) + ';')

r.append(f"""
-- Clear reversal metadata added by this migration
UPDATE public.inventory_movements
SET reversal_of_movement_id = NULL, reversed_at = NULL, reversed_by = NULL,
    reversal_reason = NULL, stocktake_ref = NULL
WHERE stocktake_ref = '{STK}';

-- 3) Restore inventory_items rows
INSERT INTO public.inventory_items AS i ({ii_cols})
SELECT {ii_cols} FROM public.inventory_items_snapshot_20260727 s
ON CONFLICT (id) DO UPDATE SET
""".format(ii_cols=', '.join(II_COLS)))
set_cols_i = [c for c in II_COLS if c != 'id']
r.append(',\n'.join(f'  {c} = EXCLUDED.{c}' for c in set_cols_i) + ';')

r.append(f"""
UPDATE public.inventory_items
SET canonical_item_id = NULL, retired_at = NULL, retired_reason = NULL
WHERE warehouse_id = '{MAIN_WH}';

-- 4) Remove stocktake session lines + header for STK
DELETE FROM public.stocktaking_lines
WHERE session_id IN (SELECT id FROM public.stocktaking_sessions
                     WHERE session_no = '{STK}' OR reference_id = '{STK}');
DELETE FROM public.stocktaking_sessions
WHERE session_no = '{STK}' OR reference_id = '{STK}';

-- 5) Drop schema objects created by the forward migration
DROP TRIGGER  IF EXISTS trg_inv_mov_reroute_canonical ON public.inventory_movements;
DROP TRIGGER  IF EXISTS trg_agouza_res_reroute        ON public.agouza_stock_reservations;
DROP FUNCTION IF EXISTS public.reroute_to_canonical();
DROP FUNCTION IF EXISTS public.reroute_res_to_canonical();
DROP VIEW     IF EXISTS public.v_inventory_balances_raw;
DROP VIEW     IF EXISTS public.v_inventory_balances;
DROP INDEX    IF EXISTS public.ux_inv_mov_idempotency_key;

ALTER TABLE public.inventory_movements
  DROP COLUMN IF EXISTS idempotency_key,
  DROP COLUMN IF EXISTS reversal_of_movement_id,
  DROP COLUMN IF EXISTS reversed_at,
  DROP COLUMN IF EXISTS reversed_by,
  DROP COLUMN IF EXISTS reversal_reason,
  DROP COLUMN IF EXISTS stocktake_ref;
ALTER TABLE public.inventory_items
  DROP COLUMN IF EXISTS canonical_item_id,
  DROP COLUMN IF EXISTS retired_at,
  DROP COLUMN IF EXISTS retired_reason;

-- 6) Verify row counts vs snapshot
DO $v$
DECLARE want bigint; got bigint;
BEGIN
  SELECT row_count INTO want FROM public.stocktake_20260727_checksums WHERE label='items_before_main';
  SELECT count(*)  INTO got  FROM public.inventory_items WHERE warehouse_id='{MAIN_WH}';
  IF want <> got THEN RAISE EXCEPTION 'Rollback: items count mismatch want=% got=%', want, got; END IF;

  SELECT row_count INTO want FROM public.stocktake_20260727_checksums WHERE label='movements_before_main';
  SELECT count(*)  INTO got  FROM public.inventory_movements WHERE warehouse_id='{MAIN_WH}';
  IF want <> got THEN RAISE EXCEPTION 'Rollback: movements count mismatch want=% got=%', want, got; END IF;
END $v$;

COMMIT;

-- Snapshot tables and checksum table remain until manually dropped:
--   DROP TABLE public.inventory_items_snapshot_20260727;
--   DROP TABLE public.inventory_movements_snapshot_20260727;
--   DROP TABLE public.stocktake_20260727_checksums;
""")

with open(RBK,'w') as f: f.write('\n'.join(r))
print('WROTE', RBK)
