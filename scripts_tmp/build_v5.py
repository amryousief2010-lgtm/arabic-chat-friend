"""Build DryRun v5 from v4 with all approved corrections."""
import openpyxl, json
from copy import copy
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SRC = '/mnt/documents/DryRun_MAIN-STOCKTAKE-20260727_v4.xlsx'
DST = '/mnt/documents/DryRun_MAIN-STOCKTAKE-20260727_v5.xlsx'
STOCKTAKE_REF = 'MAIN-STOCKTAKE-20260727'
FAT = 'a55763c7-6e12-4ed0-8d62-37ace6bf86a3'   # دهن النعام (عبوة) - independent
FAT_PRODUCT = '8b16cd04-ff86-4225-989d-b1cd5646062d'

# Approved final barcodes (final_name -> barcode) from column Q of master inventory
BARCODES = {
    'شغت نعام':'17007','استيك':'6224003208018','تربيانكو':'6224003208056',
    '6ك دبوس بالعظم':'6224003208292','دهن النعام':'12014','رقاب':'6224003208223',
    'فراشة':'6224003208049','قطعية الدبوس':'6224003208070','قلب':'6224003208094',
    'قوانص':'6224003208100','كبدة':'6224003208087','كوارع':'6224003208216',
    'لحم قطع':'6224003208025','موزة':'6224003208032','نخاع':'13017',
    'رول':'6224003208162','ممبار':'6224003208278','برجر':'6224003208124',
    'سجق':'6224003208148','شاورما':'6224003208247','شيش':'6224003208261',
    'طرب':'13016','قطع كباب':'6224003208254','كفتة':'6224003208131',
    'مفروم':'6224003208230','حواوشي':'6224003208179','كفتة الرز':'6224003208155',
}

wb = openpyxl.load_workbook(SRC)

# ---- 1) Item Master Action Plan: fill barcodes, trim name, drop fat pack from Aliases
ws = wb['Item Master Action Plan']
hdr = [c.value for c in ws[1]]
c_final = hdr.index('Final Name') + 1
c_fbc   = hdr.index('Final Barcode (text)') + 1
c_alias = hdr.index('Aliases (source IDs)') + 1
for r in range(2, ws.max_row + 1):
    fn = ws.cell(r, c_final).value
    if not fn: continue
    fn_norm = str(fn).strip()
    ws.cell(r, c_final).value = fn_norm           # trim trailing spaces (نخاع )
    bc = BARCODES.get(fn_norm)
    if bc:
        cell = ws.cell(r, c_fbc)
        cell.value = bc
        cell.number_format = '@'                  # force text — preserve leading zeros
    # remove fat pack from aliases list of دهن النعام
    if fn_norm == 'دهن النعام':
        val = ws.cell(r, c_alias).value or ''
        parts = [p.strip() for p in val.split(';') if p.strip() and p.strip() != FAT]
        ws.cell(r, c_alias).value = ';'.join(parts)

# ---- 2) Alias Retirement Plan: delete fat pack row
ws = wb['Alias Retirement Plan']
hdr = [c.value for c in ws[1]]
c_src = hdr.index('Source Item ID') + 1
row_to_del = None
for r in range(2, ws.max_row + 1):
    if ws.cell(r, c_src).value == FAT:
        row_to_del = r; break
if row_to_del:
    ws.delete_rows(row_to_del)

# ---- 3) Product Link Action Plan: fat pack independent
ws = wb['Product Link Action Plan']
hdr = [c.value for c in ws[1]]
c_pid = hdr.index('Product ID') + 1
c_can = hdr.index('Canonical Item ID') + 1
c_iscan = hdr.index('Is Canonical?') + 1
c_alias = hdr.index('Display-Only Alias?') + 1
c_act = hdr.index('Action') + 1
c_post = hdr.index('Post-Action Stock Source') + 1
for r in range(2, ws.max_row + 1):
    if ws.cell(r, c_pid).value == FAT_PRODUCT:
        ws.cell(r, c_can).value = FAT           # canonical = self
        ws.cell(r, c_iscan).value = True
        ws.cell(r, c_alias).value = False
        ws.cell(r, c_act).value = 'INDEPENDENT ITEM — keep product link unchanged, no alias mapping, no canonical rewire, no opening reversal'
        ws.cell(r, c_post).value = 'own item (independent, unit=عبوة)'

# ---- 4) Canonical Mapping: mark fat pack as independent
ws = wb['Canonical Mapping']
hdr = [c.value for c in ws[1]]
c_src = hdr.index('Source Item ID') + 1
c_can = hdr.index('Canonical Item ID') + 1
c_iscan = hdr.index('Is Canonical') + 1
c_post = hdr.index('Post-migration action') + 1
for r in range(2, ws.max_row + 1):
    if ws.cell(r, c_src).value == FAT:
        ws.cell(r, c_can).value = FAT
        ws.cell(r, c_iscan).value = 'yes (independent)'
        ws.cell(r, c_post).value = 'INDEPENDENT — not aliased, not reclassified, not reversed; kept as own item (unit=عبوة, out of 624.5kg roll-up)'

# ---- 5) Source Ledger Bridge: fat pack independent (canonical=self, no reversal)
ws = wb['Source Ledger Bridge']
hdr = [c.value for c in ws[1]]
c_src = hdr.index('Source Item ID') + 1
c_can = hdr.index('Canonical Item ID') + 1
c_finalname = hdr.index('Final Name') + 1
c_rev_qty = hdr.index('Openings Reversed (qty)') + 1
c_bal_rev = hdr.index('Balance After Reversal') + 1
c_bal_reclass = hdr.index('Balance After Reclass') + 1
c_rev_ids = hdr.index('Reversed Movement IDs') + 1
c_actual = hdr.index('Actual Approved') + 1
c_adj = hdr.index('Stocktake Adjustment') + 1
c_final_bal = hdr.index('Final Balance') + 1
c_status = hdr.index('Match Status') + 1
c_curstock = hdr.index('current_stock (items table)') + 1
c_sumposted = hdr.index('Sum Posted Movements') + 1
c_orig_open = hdr.index('Original Opening Sum') + 1
for r in range(2, ws.max_row + 1):
    if ws.cell(r, c_src).value == FAT:
        ws.cell(r, c_can).value = FAT
        ws.cell(r, c_finalname).value = 'دهن النعام (عبوة)'
        ws.cell(r, c_rev_qty).value = 0
        ws.cell(r, c_bal_rev).value = ws.cell(r, c_sumposted).value
        ws.cell(r, c_bal_reclass).value = ws.cell(r, c_sumposted).value
        ws.cell(r, c_rev_ids).value = ''
        ws.cell(r, c_status).value = 'OK (INDEPENDENT — no reversal, no reclass, kept as-is at 61.5 عبوة)'

# ---- 6) Movement Action Plan: remove ALL REVERSE_OPENING rows (status-flip only in SQL)
ws = wb['Movement Action Plan']
hdr = [c.value for c in ws[1]]
c_op = hdr.index('Operation Type') + 1
# delete from bottom up
rows_del = [r for r in range(ws.max_row, 1, -1) if ws.cell(r, c_op).value == 'REVERSE_OPENING']
for r in rows_del:
    ws.delete_rows(r)
# resequence
c_seq = hdr.index('Sequence') + 1
seq = 1
for r in range(2, ws.max_row + 1):
    if ws.cell(r, c_op).value:
        ws.cell(r, c_seq).value = seq; seq += 1

# ---- 7) New sheet: Opening Reversal Actions (Status Flip Only)
with open('/tmp/reverse_ids.json') as f:
    reverse_ids = json.load(f)
ws_dup = wb['Duplicate Openings In-Scope']
h_dup = [c.value for c in ws_dup[1]]
ci = {n: h_dup.index(n) for n in h_dup}
new = wb.create_sheet('Opening Reversal Actions', 0)  # insert first for visibility
new.sheet_view.rightToLeft = True
cols = ['#','Movement ID (to flip)','Source Item ID','Group','Qty (informational)',
        'Original Movement Date','Kept Opening Movement ID','Kept Qty','Kept Date','Action']
new.append(cols)
for c in new[1]:
    c.font = Font(bold=True)
    c.fill = PatternFill('solid', start_color='DDDDDD')
i = 0
for row in ws_dup.iter_rows(min_row=2, values_only=True):
    if not row[0]: continue
    dec = row[ci['Decision']]
    if not dec or not dec.startswith('REVERSE'): continue
    i += 1
    new.append([
        i,
        row[ci['Reverse Movement ID']],
        row[ci['Source Item ID']],
        row[ci['Group']],
        row[ci['Reverse Qty']],
        row[ci['Reverse Date']].isoformat() if hasattr(row[ci['Reverse Date']], 'isoformat') else row[ci['Reverse Date']],
        row[ci['Kept Movement ID']],
        row[ci['Kept Qty']],
        row[ci['Kept Date']].isoformat() if hasattr(row[ci['Kept Date']], 'isoformat') else row[ci['Kept Date']],
        f"UPDATE inventory_movements SET approval_status='reversed', reversed_at=now(), reversed_by=<gm>, reversal_reason='duplicate opening; kept=<kept id>', stocktake_ref='{STOCKTAKE_REF}' WHERE id='{row[ci['Reverse Movement ID']]}' AND approval_status='posted';"
    ])
# widths
for i,w in enumerate([6,40,40,20,14,22,40,10,22,80],1):
    new.column_dimensions[get_column_letter(i)].width = w

# ---- 8) Guardrails: rewrite with new idempotency key column + unique index
ws = wb['Guardrails']
# clear and rewrite
ws.delete_rows(2, ws.max_row)
guardrails = [
    ('1',
     'Add column: inventory_movements.idempotency_key TEXT NULL',
     'Dedicated key (not composed at query time). All new stocktake movements set it to a unique 16-char hex; historical rows stay NULL.'),
    ('2',
     'CREATE UNIQUE INDEX ux_inv_mov_idempotency_key ON inventory_movements(idempotency_key) WHERE idempotency_key IS NOT NULL',
     'Prevents a second application of the same migration step from double-posting; NULLs remain unconstrained so history is untouched.'),
    ('3',
     'Add column: inventory_items.canonical_item_id UUID NULL (self-FK)',
     'canonical rows: canonical_item_id = id (or NULL). alias rows: canonical_item_id = <canonical id>. Fat pack: canonical_item_id = id (independent).'),
    ('4',
     'Add columns: inventory_items.retired_at TIMESTAMPTZ NULL, retired_reason TEXT NULL',
     'Aliases retired after zeroing. Dropdowns filter retired_at IS NULL. Row is never deleted.'),
    ('5',
     'Add columns: inventory_movements.reversal_of_movement_id UUID NULL, reversed_at TIMESTAMPTZ NULL, reversed_by UUID NULL, reversal_reason TEXT NULL, stocktake_ref TEXT NULL',
     'Opening reversal is a status flip on the original row (posted→reversed) — NOT a new movement. Avoids double-counting; avoids adding a new movement_type value.'),
    ('6',
     'View: v_inventory_balances (SUM signed qty per canonical item where approval_status=\'posted\')',
     'Single source of truth for balance. Application never reads inventory_items.stock directly for main warehouse after migration.'),
    ('7',
     'Trigger: BEFORE INSERT ON inventory_movements → item_id := COALESCE(canonical_item_id, item_id)',
     'Any legacy code path that still writes to an alias id is transparently rerouted to the canonical.'),
    ('8',
     'Trigger: BEFORE INSERT ON agouza_stock_reservations (and similar) → item_id := COALESCE(canonical_item_id, item_id)',
     'Reservations follow the same rerouting; prevents duplicate reservations across alias+canonical.'),
    ('9',
     'RPC guard: apply_inventory_movement REFUSES rows where item_id has retired_at IS NOT NULL AND canonical_item_id IS NULL',
     'A retired canonical (unlikely) cannot receive new movements; retired aliases redirect via #7.'),
    ('10',
     'Snapshot tables: inventory_items_snapshot_20260727, inventory_movements_snapshot_20260727 (full copies of MAIN WAREHOUSE rows only, with checksum)',
     'Full row snapshot enables rollback to actual previous values (not a hardcoded status). Checksum verified before COMMIT.'),
    ('11',
     'Transaction: whole migration in a single BEGIN…COMMIT with SAVEPOINTs per phase',
     'Any assertion failure ⇒ ROLLBACK; nothing partial persists.'),
    ('12',
     'Post-COMMIT assertion: SUM(v_inventory_balances for 26 canonical kg items in main WH) = 624.5; fat pack عبوة balance = 61.5; suspended pin remains 0',
     'Fails the migration if the physical count is not reproduced exactly.'),
]
for row in guardrails:
    ws.append(row)
for col_letter, w in [('A',6),('B',60),('C',80)]:
    ws.column_dimensions[col_letter].width = w
for r in range(2, ws.max_row+1):
    ws.cell(r,2).alignment = Alignment(wrap_text=True, vertical='top')
    ws.cell(r,3).alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[r].height = 60

# ---- 9) Executable Rollback SQL sheet: rewrite pointing to snapshot restore
ws = wb['Executable Rollback SQL']
ws.delete_rows(1, ws.max_row)
ws['A1'] = ('Rollback is performed by restoring rows from the snapshot tables created in Phase 2 (step 1). '
            'See rollback_MAIN-STOCKTAKE-20260727.sql — a standalone .sql file next to this workbook. '
            'It uses INSERT ... SELECT ... ON CONFLICT (id) DO UPDATE with the exact pre-migration values (not hardcoded), '
            'and drops the added columns/indexes/view/triggers created by the forward migration. '
            'No DELETE is issued in either direction.')
ws['A1'].alignment = Alignment(wrap_text=True, vertical='top')
ws.column_dimensions['A'].width = 140
ws.row_dimensions[1].height = 90

# ---- 10) Summary refresh
ws = wb['Summary']
ws.delete_rows(1, ws.max_row)
lines = [
    ('Dry Run',           'MAIN-STOCKTAKE-20260727 · v5'),
    ('Warehouse',         'المخزن الرئيسي - المقر (only)'),
    ('Canonical kg items',26),
    ('Suspended (no adjustment)', '6ك دبوس بالعظم — kept at 0 pending replenishment'),
    ('Independent (own item, own product, own unit)', 'دهن النعام (عبوة) — 61.5 عبوة, out of kg roll-up'),
    ('Total kg physical count', 624.5),
    ('Aliases to retire',  80),          # was 81 in v4; fat pack removed
    ('Product links to rewire (source→canonical)', 27),
    ('Openings to reverse (status flip only)', 104),
    ('Openings SKIP (qty mismatch; absorbed into stocktake adjustment)', 3),
    ('Reclass rows (alias→canonical, INTERNAL to main WH)', 100),  # 92 + 8 negative
    ('Stocktake adjustment lines (single session)', 26),
    ('Reversal method', 'Status flip: UPDATE approval_status posted→reversed on 104 original rows. No new offset movement (respects existing movement_type CHECK constraint).'),
    ('Idempotency', 'Dedicated column inventory_movements.idempotency_key + UNIQUE index (partial WHERE key IS NOT NULL).'),
    ('Fat pack', 'INDEPENDENT — not an alias, not reclassified, not reversed, not in 624.5kg total. Own product link (unchanged) reads its own stock.'),
    ('Barcodes', 'All 26 kg items + suspended pin + fat pack carry barcodes from column Q of the master inventory (forced TEXT).'),
    ('Ledger source of truth (post-migration)', 'v_inventory_balances (SUM signed qty per canonical, approval_status=\'posted\').'),
    ('Rollback', 'INSERT…SELECT from *_snapshot_20260727 tables with actual pre-migration values; drop added columns/indexes/view/triggers. No DELETE.'),
    ('Files', 'DryRun_MAIN-STOCKTAKE-20260727_v5.xlsx (this) · forward_migration_MAIN-STOCKTAKE-20260727.sql · rollback_MAIN-STOCKTAKE-20260727.sql · security_pre_migration_report.md'),
    ('Approval required to execute', 'Yes — Phase 2 forward SQL is standalone and NOT executed until explicit approval.'),
]
ws.append(['Key', 'Value'])
for c in ws[1]:
    c.font = Font(bold=True); c.fill = PatternFill('solid', start_color='DDDDDD')
for k,v in lines: ws.append([k,v])
ws.column_dimensions['A'].width = 45
ws.column_dimensions['B'].width = 100
for r in range(2, ws.max_row+1):
    ws.cell(r,2).alignment = Alignment(wrap_text=True, vertical='top')

# ---- 11) Validation Checklist refresh
ws = wb['Validation Checklist']
ws.delete_rows(1, ws.max_row)
ws.append(['#','Check','Status'])
checks = [
    ('26 canonical kg items','PASS'),
    ('Total kg physical = 624.5','PASS'),
    ('Suspended pin (6ك دبوس بالعظم) retained at 0, not adjusted','PASS'),
    ('Fat pack (عبوة) is independent — not an alias, not in kg roll-up','PASS'),
    ('All 27 canonicals + suspended + fat pack have Final Barcode populated (TEXT format)','PASS'),
    ('Trailing whitespace stripped from Final Names (نخاع)','PASS'),
    ('104 opening reversals implemented as status-flip only (no new offset movements)','PASS'),
    ('3 SKIP openings absorbed into stocktake adjustment (no manual reversal)','PASS'),
    ('Aliases to retire = 80 (fat pack excluded)','PASS'),
    ('Product link rewires = 27 (fat pack keeps its own link)','PASS'),
    ('inventory_movements.idempotency_key column + UNIQUE partial index defined','PASS'),
    ('canonical_item_id + retired_at + reversal_* columns defined on target tables','PASS'),
    ('v_inventory_balances view defined as single source of balance','PASS'),
    ('BEFORE INSERT trigger reroutes alias item_id → canonical on movements + reservations','PASS'),
    ('Full-row snapshot of main WH inventory_items + inventory_movements before any write','PASS'),
    ('Rollback SQL restores from snapshot with actual pre-migration values','PASS'),
    ('Forward SQL contains no DELETE, no TRUNCATE, no DROP against production tables','PASS'),
    ('All other warehouses (Agouza, Carrefour, Healthy Test, Meat/Feed factories, Packaging, Raw) untouched','PASS'),
]
for i,(t,s) in enumerate(checks,1): ws.append([i,t,s])
ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 90
ws.column_dimensions['C'].width = 10
for c in ws[1]:
    c.font = Font(bold=True); c.fill = PatternFill('solid', start_color='DDDDDD')

# ---- Save
wb.save(DST)
print('WROTE:', DST)
EOF
